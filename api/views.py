from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from django.shortcuts import render
from django.db.models import Count, Q
import json
import base64
from datetime import datetime
from django.core.files.base import ContentFile
from .models import Planilla, Mascota, Responsable, Veterinario, RegistroPerdidas
from .serializers import PlanillaSerializer, MascotaSerializer, ResponsableSerializer
from django.shortcuts import render, redirect, get_object_or_404
from .forms import ResponsableForm, MascotaFormSet
from django.contrib.auth.decorators import login_required
@login_required
def elegir_planilla(request):
    """
    Muestra las planillas asignadas al veterinario logueado
    para que elija en cuál va a agregar responsables y mascotas.
    """
    planillas = Planilla.objects.filter(assigned_to=request.user)
    return render(request, 'api/elegir_planilla.html', {
        'planillas': planillas
    })

def crear_responsable_con_mascotas(request, planilla_id):
    # 1) Obtén la planilla o 404
    planilla = get_object_or_404(Planilla, id=planilla_id)

    if request.method == 'POST':
        form = ResponsableForm(request.POST)
        formset = MascotaFormSet(request.POST, request.FILES)

        if form.is_valid() and formset.is_valid():
            # 2) Guarda Responsable
            responsable = form.save(commit=False)
            responsable.planilla = planilla
            responsable.save()

            # 3) Asocia el formset al responsable
            formset.instance = responsable
            formset.save()

            return redirect('detalle_planilla', pk=planilla.id)  # o a donde quieras
    else:
        form = ResponsableForm()
        formset = MascotaFormSet()

    return render(request, 'api/responsable_form.html', {
        'form': form,
        'formset': formset,
        'planilla': planilla
    })

class ResponsableViewSet(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request, planilla_id=None):
        """GET: Lista responsables de una planilla específica"""
        if planilla_id:
            responsables = Responsable.objects.filter(planilla_id=planilla_id)
            # Filtrar por created_by si es vacunador
            if request.user.is_authenticated and hasattr(request.user, 'tipo_usuario') and request.user.tipo_usuario == 'vacunador':
                responsables = responsables.filter(created_by=request.user)
        else:
            responsables = Responsable.objects.all()
            # Filtrar por created_by si es vacunador
            if request.user.is_authenticated and hasattr(request.user, 'tipo_usuario') and request.user.tipo_usuario == 'vacunador':
                responsables = responsables.filter(created_by=request.user)
        return Response(ResponsableSerializer(responsables, many=True).data)
    
    def post(self, request, planilla_id=None):
        """POST: Crea un nuevo responsable con sus mascotas"""
        try:
            if planilla_id:
                planilla = Planilla.objects.get(id=planilla_id)
            else:
                planilla_id = request.data.get('planilla_id')
                planilla = Planilla.objects.get(id=planilla_id)
        except Planilla.DoesNotExist:
            return Response({'error': 'Planilla no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        # Crear responsable
        responsable_data = {
            'nombre': request.data.get('nombre'),
            'telefono': request.data.get('telefono'),
            'finca': request.data.get('finca'),
            'zona': request.data.get('zona', 'Sin especificar'),
            'nombre_zona': request.data.get('nombre_zona', 'Sin especificar'),
            'lote_vacuna': request.data.get('lote_vacuna', 'Sin especificar'),
            'planilla': planilla.id
        }

        responsable_serializer = ResponsableSerializer(data=responsable_data)
        if responsable_serializer.is_valid():
            # Asignar created_by si el usuario está autenticado
            if request.user.is_authenticated:
                print(f"Usuario autenticado: {request.user.username} (ID: {request.user.id})")
                responsable = responsable_serializer.save(created_by=request.user)
                print(f"Responsable creado con created_by: {responsable.created_by}")
            else:
                print("Usuario no autenticado - responsable sin created_by")
                responsable = responsable_serializer.save()

            # Crear mascotas si se proporcionan
            mascotas_data = request.data.get('mascotas', [])
            mascotas_creadas = []
            mascotas_validas = 0
            errores_mascotas = []

            print(f"DEBUG: mascotas_data tipo: {type(mascotas_data)}")
            print(f"DEBUG: mascotas_data contenido (primeros 200 chars): {str(mascotas_data)[:200]}")

            # Si mascotas_data es un string, parsearlo como JSON
            if isinstance(mascotas_data, str):
                try:
                    mascotas_data = json.loads(mascotas_data)
                    print(f"JSON parseado exitosamente. Tipo: {type(mascotas_data)}, Items: {len(mascotas_data)}")
                except json.JSONDecodeError:
                    print(f"Error parseando JSON de mascotas: {mascotas_data}")
                    mascotas_data = []

            # Procesar cada mascota
            for i, mascota_data in enumerate(mascotas_data):
                print(f"Procesando mascota {i+1}: tipo {type(mascota_data)}")

                # Verificar si mascota_data es un diccionario
                if isinstance(mascota_data, dict):
                    # Hacer una copia del diccionario para evitar el error
                    mascota_data_copy = mascota_data.copy()
                    mascota_data_copy['responsable'] = responsable.id

                    # Validar que tenga nombre y color (campos obligatorios solo si no están vacíos)
                    nombre = mascota_data_copy.get('nombre', '').strip()
                    color = mascota_data_copy.get('color', '').strip()

                    # Si tanto nombre como color están vacíos, saltamos esta mascota (no es un error)
                    if not nombre and not color:
                        print(f"Mascota {i+1} saltada: campos vacíos, probablemente slot no usado")
                        continue

                    # Si tiene nombre pero no color, o viceversa, entonces sí es error
                    if not nombre:
                        print(f"Mascota {i+1} saltada: sin nombre válido pero tiene otros datos")
                        errores_mascotas.append(f"Mascota {i+1}: el nombre es obligatorio cuando se llena el formulario")
                        continue

                    if not color:
                        print(f"Mascota {i+1} saltada: sin color válido pero tiene otros datos")
                        errores_mascotas.append(f"Mascota {i+1}: el color es obligatorio cuando se llena el formulario")
                        continue

                    # DEBUG: Verificar foto
                    foto_base64 = mascota_data_copy.get('foto')
                    foto_index = mascota_data_copy.get('foto_index')
                    print(f"Foto encontrada: {type(foto_base64)}, longitud: {len(str(foto_base64)) if foto_base64 else 0}")
                    print(f"Foto_index encontrado: {foto_index}")

                    # Procesar foto si existe
                    if foto_base64 and isinstance(foto_base64, str) and len(foto_base64) > 100:
                        try:
                            print("Procesando foto base64...")

                            # Remover el prefijo data:image si existe
                            if foto_base64.startswith('data:image'):
                                foto_base64 = foto_base64.split(',')[1]
                                print("Prefijo data:image removido")

                            # Decodificar base64
                            foto_data = base64.b64decode(foto_base64)
                            foto_file = ContentFile(foto_data, name=f'mascota_{mascota_data_copy.get("nombre", "sin_nombre")}.png')
                            mascota_data_copy['foto'] = foto_file
                            print(f"Foto procesada para mascota: {mascota_data_copy.get('nombre')}")
                        except Exception as e:
                            print(f"Error procesando foto: {e}")
                            mascota_data_copy.pop('foto', None)
                    else:
                        print(f"Foto no valida o muy corta: {len(str(foto_base64)) if foto_base64 else 0} chars")
                        # Limpiar campos de foto para evitar errores
                        mascota_data_copy.pop('foto', None)
                        mascota_data_copy.pop('foto_index', None)

                    print(f"Datos de mascota antes del serializer: {mascota_data_copy}")
                    mascota_serializer = MascotaSerializer(data=mascota_data_copy)
                    if mascota_serializer.is_valid():
                        # Asignar created_by si el usuario está autenticado
                        if request.user.is_authenticated:
                            mascota = mascota_serializer.save(created_by=request.user)
                            print(f"Mascota creada con created_by: {mascota.created_by}")
                        else:
                            print("Usuario no autenticado - mascota sin created_by")
                            mascota = mascota_serializer.save()
                        serialized_data = mascota_serializer.data
                        print(f"Datos serializados de mascota: {serialized_data}")
                        mascotas_creadas.append(serialized_data)
                        mascotas_validas += 1
                        print(f"Mascota {mascota.nombre} creada exitosamente")
                    else:
                        print(f"Error en mascota serializer: {mascota_serializer.errors}")
                        errores_mascotas.append(f"Mascota {i+1}: {mascota_serializer.errors}")
                else:
                    print(f"mascota_data no es un diccionario: {type(mascota_data)}, valor: {mascota_data}")
                    errores_mascotas.append(f"Mascota {i+1}: formato de datos inválido")

            # Verificar que se haya creado al menos una mascota válida
            if mascotas_validas == 0 and len(mascotas_data) > 0:
                # Si se enviaron mascotas pero ninguna es válida, eliminar el responsable
                responsable.delete()
                return Response({
                    'error': 'Se requiere al menos una mascota válida con nombre y color obligatorios',
                    'errores_mascotas': errores_mascotas
                }, status=status.HTTP_400_BAD_REQUEST)

            response_data = responsable_serializer.data
            response_data['mascotas'] = mascotas_creadas
            response_data['mascotas_creadas'] = mascotas_validas
            response_data['errores_mascotas'] = errores_mascotas

            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            return Response(responsable_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MascotaViewSet(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request, responsable_id):
        """GET: Lista mascotas de un responsable específico"""
        try:
            # Filtrar mascotas por responsable y por created_by si es vacunador
            mascotas = Mascota.objects.filter(responsable_id=responsable_id)
            if request.user.is_authenticated and hasattr(request.user, 'tipo_usuario') and request.user.tipo_usuario == 'vacunador':
                mascotas = mascotas.filter(created_by=request.user)
            return Response(MascotaSerializer(mascotas, many=True).data)
        except Responsable.DoesNotExist:
            return Response({'error': 'Responsable no encontrado'}, status=status.HTTP_404_NOT_FOUND)
    
    def post(self, request, responsable_id):
        """POST: Crea una nueva mascota para un responsable"""
        try:
            responsable = Responsable.objects.get(id=responsable_id)
        except Responsable.DoesNotExist:
            return Response({'error': 'Responsable no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        mascota_data = request.data.copy()
        mascota_data['responsable'] = responsable_id
        
        serializer = MascotaSerializer(data=mascota_data)
        if serializer.is_valid():
            # Asignar created_by si el usuario está autenticado
            if request.user.is_authenticated:
                mascota = serializer.save(created_by=request.user)
            else:
                mascota = serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mis_planillas(request):
    """Lista planillas filtradas por el usuario.

    Reglas:
    - Si viene ?usuario=<username>, filtra por ese username (útil para apps móviles).
    - Si viene ?municipio=<municipio>, filtra por ese municipio.
    - En otro caso, filtra por el usuario autenticado (request.user).
    - Incluye planillas donde el usuario es asignado principal O adicional.
    """
    username = request.query_params.get('usuario') or request.user.username
    municipio = request.query_params.get('municipio')

    qs = Planilla.objects.all()

    # Filtrar por municipio si se proporciona
    if municipio:
        qs = qs.filter(municipio__icontains=municipio)

    # Filtrar por usuario
    if username and not municipio:  # Solo filtrar por usuario si no estamos filtrando por municipio
        # Filtrar por asignaciones principales Y adicionales
        qs = qs.filter(
            Q(assigned_to__username=username) |  # Vacunador principal (compatibilidad app móvil)
            Q(vacunadores_adicionales__username=username) |  # Vacunadores adicionales
            Q(tecnico_asignado__username=username) |  # Técnico principal
            Q(tecnicos_adicionales__username=username)  # Técnicos adicionales
        ).distinct()

    return Response(PlanillaSerializer(qs, many=True, context={'request': request}).data)


@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def mascotas_planilla(request, pk):
    try:
        plan = Planilla.objects.get(pk=pk)
    except Planilla.DoesNotExist:
        return Response(status=404)

    if request.method == 'GET':
        return Response(MascotaSerializer(plan.mascotas.all(), many=True).data)

    # POST → crea mascota
    serializer = MascotaSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save(planilla=plan)
    return Response(serializer.data, status=201)


def landing_page(request):
    """Vista para la landing page con estadísticas generales"""
    # Calcular estadísticas generales
    total_planillas = Planilla.objects.count()
    total_responsables = Responsable.objects.count()
    total_mascotas = Mascota.objects.count()
    # Todos están vacunados, el campo se refiere a tarjeta de vacunación previa
    mascotas_con_tarjeta_previa = Mascota.objects.filter(antecedente_vacunal=True).count()
    
    context = {
        'total_planillas': total_planillas,
        'total_responsables': total_responsables,
        'total_mascotas': total_mascotas,
        'mascotas_vacunadas': total_mascotas,  # Todos están vacunados
        'mascotas_con_tarjeta_previa': mascotas_con_tarjeta_previa,
    }
    
    return render(request, 'api/landing.html', context)


@login_required
def reportes_view(request):
    """Vista para reportes detallados por municipio - filtrada por rol de usuario"""
    user = request.user
    
    # Filtrar planillas según el rol del usuario
    if user.tipo_usuario == 'administrador':
        # Administradores ven todas las planillas
        planillas = Planilla.objects.select_related().prefetch_related('responsables__mascotas')
    elif user.tipo_usuario == 'tecnico':
        # Técnicos ven solo las planillas de sus municipios asignados
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).select_related().prefetch_related('responsables__mascotas').distinct()
    elif user.tipo_usuario == 'vacunador':
        # Vacunadores ven solo sus planillas asignadas
        planillas = Planilla.objects.filter(
            Q(assigned_to=user) |
            Q(vacunadores_adicionales=user)
        ).select_related().prefetch_related('responsables__mascotas').distinct()
    else:
        planillas = Planilla.objects.none()
    
    # Diccionario para agrupar por municipio
    municipios_stats = {}
    
    for planilla in planillas:
        municipio = planilla.municipio
        if municipio not in municipios_stats:
            municipios_stats[municipio] = {
                'municipio': municipio,
                'total_mascotas': 0,
                'con_tarjeta_previa': 0,
                'sin_tarjeta_previa': 0,
                'zona_urbana': 0,
                'zona_rural': 0,
                'perros': 0,
                'gatos': 0,
                'responsables': 0,
                'planillas': 0,
            }
        
        # Contar planillas por municipio
        municipios_stats[municipio]['planillas'] += 1
        
        # Contar responsables y mascotas por planilla
        # Filtrar responsables según el rol del usuario
        if user.tipo_usuario == 'vacunador':
            # Vacunadores solo ven sus propios registros
            responsables = planilla.responsables.filter(created_by=user)
        else:
            # Técnicos y administradores ven todos los registros de la planilla
            responsables = planilla.responsables.all()
            
        for responsable in responsables:
            municipios_stats[municipio]['responsables'] += 1
            
            # Filtrar mascotas según el rol del usuario
            if user.tipo_usuario == 'vacunador':
                mascotas = responsable.mascotas.filter(created_by=user)
            else:
                mascotas = responsable.mascotas.all()
            
            for mascota in mascotas:
                municipios_stats[municipio]['total_mascotas'] += 1
                
                # Contar por tarjeta de vacunación previa (todos son vacunados ahora)
                if mascota.antecedente_vacunal:
                    municipios_stats[municipio]['con_tarjeta_previa'] += 1
                else:
                    municipios_stats[municipio]['sin_tarjeta_previa'] += 1
                
                # Contar por zona (basado en la zona del responsable)
                # Mapear zona del responsable a urbano/rural
                if responsable.zona == 'barrio':
                    municipios_stats[municipio]['zona_urbana'] += 1
                elif responsable.zona == 'vereda':
                    municipios_stats[municipio]['zona_rural'] += 1
                elif responsable.zona == 'centro poblado':
                    # Centro poblado puede ser urbano o rural, por defecto urbano
                    municipios_stats[municipio]['zona_urbana'] += 1
                else:
                    # Fallback: usar el valor de la planilla si la zona del responsable no está definida
                    if planilla.urbano_rural == 'urbano':
                        municipios_stats[municipio]['zona_urbana'] += 1
                    else:
                        municipios_stats[municipio]['zona_rural'] += 1
                
                # Contar por tipo
                if mascota.tipo == 'perro':
                    municipios_stats[municipio]['perros'] += 1
                else:
                    municipios_stats[municipio]['gatos'] += 1
    
    # Calcular porcentajes de tarjetas previas
    reportes_municipio = []
    for stats in municipios_stats.values():
        if stats['total_mascotas'] > 0:
            porcentaje_tarjeta_previa = round((stats['con_tarjeta_previa'] / stats['total_mascotas']) * 100, 1)
        else:
            porcentaje_tarjeta_previa = 0
        
        stats['porcentaje_tarjeta_previa'] = porcentaje_tarjeta_previa
        reportes_municipio.append(stats)
    
    # Ordenar por municipio
    reportes_municipio.sort(key=lambda x: x['municipio'])
    
    # Para administradores, generar datos detallados por municipio
    reportes_detallados_municipio = {}
    if user.tipo_usuario == 'administrador':
        for planilla in planillas:
            municipio = planilla.municipio
            if municipio not in reportes_detallados_municipio:
                reportes_detallados_municipio[municipio] = {
                    'municipio': municipio,
                    'responsables': []
                }
            
            # Obtener todos los responsables de esta planilla
            responsables = planilla.responsables.all()
            for responsable in responsables:
                mascotas = responsable.mascotas.all()
                responsable_data = {
                    'nombre': responsable.nombre,
                    'telefono': responsable.telefono,
                    'finca': responsable.finca,
                    'zona': responsable.zona,
                    'nombre_zona': responsable.nombre_zona,
                    'lote_vacuna': responsable.lote_vacuna,
                    'creado': responsable.creado,
                    'created_by': responsable.created_by.username if responsable.created_by else 'N/A',
                    'mascotas': []
                }
                
                for mascota in mascotas:
                    mascota_data = {
                        'nombre': mascota.nombre,
                        'tipo': mascota.tipo,
                        'raza': mascota.raza,
                        'color': mascota.color,
                        'antecedente_vacunal': mascota.antecedente_vacunal,
                        'creado': mascota.creado,
                        'created_by': mascota.created_by.username if mascota.created_by else 'N/A',
                    }
                    responsable_data['mascotas'].append(mascota_data)
                
                reportes_detallados_municipio[municipio]['responsables'].append(responsable_data)
        
        # Convertir a lista y ordenar
        reportes_detallados_municipio = list(reportes_detallados_municipio.values())
        reportes_detallados_municipio.sort(key=lambda x: x['municipio'])
    
    # Calcular estadísticas generales
    total_municipios = len(municipios_stats)
    total_planillas = sum(stats['planillas'] for stats in municipios_stats.values())
    total_responsables = sum(stats['responsables'] for stats in municipios_stats.values())
    total_mascotas = sum(stats['total_mascotas'] for stats in municipios_stats.values())
    total_con_tarjeta = sum(stats['con_tarjeta_previa'] for stats in municipios_stats.values())
    total_sin_tarjeta = sum(stats['sin_tarjeta_previa'] for stats in municipios_stats.values())
    total_urbano = sum(stats['zona_urbana'] for stats in municipios_stats.values())
    total_rural = sum(stats['zona_rural'] for stats in municipios_stats.values())
    total_perros = sum(stats['perros'] for stats in municipios_stats.values())
    total_gatos = sum(stats['gatos'] for stats in municipios_stats.values())
    
    # Calcular porcentajes generales
    if total_mascotas > 0:
        porcentaje_general_tarjeta = round((total_con_tarjeta / total_mascotas) * 100, 1)
        porcentaje_urbano = round((total_urbano / total_mascotas) * 100, 1)
        porcentaje_rural = round((total_rural / total_mascotas) * 100, 1)
        porcentaje_perros = round((total_perros / total_mascotas) * 100, 1)
        porcentaje_gatos = round((total_gatos / total_mascotas) * 100, 1)
    else:
        porcentaje_general_tarjeta = 0
        porcentaje_urbano = 0
        porcentaje_rural = 0
        porcentaje_perros = 0
        porcentaje_gatos = 0
    
    context = {
        'reportes_municipio': reportes_municipio,
        'reportes_detallados_municipio': reportes_detallados_municipio if user.tipo_usuario == 'administrador' else [],
        'total_municipios': total_municipios,
        'total_planillas': total_planillas,
        'total_responsables': total_responsables,
        'total_mascotas': total_mascotas,
        'total_con_tarjeta': total_con_tarjeta,
        'total_sin_tarjeta': total_sin_tarjeta,
        'total_urbano': total_urbano,
        'total_rural': total_rural,
        'total_perros': total_perros,
        'total_gatos': total_gatos,
        'porcentaje_general_tarjeta': porcentaje_general_tarjeta,
        'porcentaje_urbano': porcentaje_urbano,
        'porcentaje_rural': porcentaje_rural,
        'porcentaje_perros': porcentaje_perros,
        'porcentaje_gatos': porcentaje_gatos,
    }
    
    return render(request, 'api/reportes.html', context)


from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.shortcuts import redirect

def login_view(request):
    """Vista para el inicio de sesión"""
    if request.user.is_authenticated:
        return redirect('dashboard_principal')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'¡Bienvenido, {user.username}!')
            return redirect('dashboard_principal')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    
    return render(request, 'api/login.html')


@login_required
def dashboard_principal(request):
    """Dashboard principal que redirije según el tipo de usuario"""
    user = request.user
    
    if user.tipo_usuario == 'administrador':
        # Administradores van a la landing page con estadísticas completas
        return redirect('landing')
    elif user.tipo_usuario == 'tecnico':
        # Técnicos van directamente a reportes (pueden ver todos los registros de sus municipios)
        return redirect('reportes')
    elif user.tipo_usuario == 'vacunador':
        # Vacunadores van a reportes (solo ven sus propios registros)
        return redirect('reportes')
    else:
        messages.error(request, 'Tipo de usuario no válido.')
        return redirect('login')


def logout_view(request):
    """Vista para cerrar sesión"""
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('login')


# ========== NUEVAS VISTAS CON PERMISOS POR TIPO DE USUARIO ==========

@login_required
def dashboard_vacunador(request):
    """Dashboard para vacunadores - Solo ven sus propias planillas, responsables y mascotas"""
    if request.user.tipo_usuario != 'vacunador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Planillas donde el usuario es vacunador principal O adicional
    planillas = Planilla.objects.filter(
        Q(assigned_to=request.user) |  # Vacunador principal
        Q(vacunadores_adicionales=request.user)  # Vacunador adicional
    ).distinct()

    # Solo mascotas creadas por este vacunador
    mascotas = Mascota.objects.filter(created_by=request.user)

    # Solo responsables creados por este vacunador
    responsables = Responsable.objects.filter(created_by=request.user)

    # Calcular estadísticas de zona basado en mascotas del vacunador
    total_urbano = 0
    total_rural = 0

    for mascota in mascotas:
        responsable = mascota.responsable
        # Mapear zona del responsable a urbano/rural
        if responsable.zona == 'barrio':
            total_urbano += 1
        elif responsable.zona == 'vereda':
            total_rural += 1
        elif responsable.zona == 'centro poblado':
            # Centro poblado puede ser urbano o rural, por defecto urbano
            total_urbano += 1
        else:
            # Fallback: usar el valor de la planilla si la zona del responsable no está definida
            if responsable.planilla.urbano_rural == 'urbano':
                total_urbano += 1
            else:
                total_rural += 1

    # Calcular porcentajes
    total_mascotas_count = mascotas.count()
    if total_mascotas_count > 0:
        porcentaje_urbano = round((total_urbano / total_mascotas_count) * 100, 1)
        porcentaje_rural = round((total_rural / total_mascotas_count) * 100, 1)
    else:
        porcentaje_urbano = 0
        porcentaje_rural = 0

    # Estadísticas de mascotas por día para el vacunador
    from django.db.models import Count, Min
    from datetime import datetime, timedelta, date

    # Obtener la fecha más antigua de mascotas registradas por este vacunador
    primera_mascota = mascotas.aggregate(Min('creado'))['creado__min']
    if primera_mascota:
        fecha_inicio = primera_mascota.date() if hasattr(primera_mascota, 'date') else primera_mascota
    else:
        fecha_inicio = datetime.now().date()

    fecha_fin = datetime.now().date()

    # Crear lista de todos los días desde el inicio hasta hoy
    dias_completos = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        dias_completos.append(fecha_actual)
        fecha_actual += timedelta(days=1)

    # Obtener datos de mascotas por día para este vacunador
    mascotas_query = mascotas.extra(
        select={'dia': 'date(api_mascota.creado)'}
    ).values('dia').annotate(cantidad=Count('id'))

    # Crear diccionario con los datos
    mascotas_dict = {item['dia']: item['cantidad'] for item in mascotas_query}

    # Crear lista completa con todos los días
    mascotas_por_dia = []
    for dia in dias_completos:
        mascotas_por_dia.append({
            'dia': dia,
            'cantidad': mascotas_dict.get(str(dia), 0)
        })

    # Obtener municipios únicos para el selector
    municipios_unicos = list(planillas.values_list('municipio', flat=True).distinct().order_by('municipio'))

    context = {
        'user_type': 'Vacunador',
        'planillas': planillas,
        'municipios_unicos': municipios_unicos,
        'total_planillas': planillas.count(),
        'total_responsables': responsables.count(),
        'total_mascotas': mascotas.count(),
        'mascotas_con_tarjeta': mascotas.filter(antecedente_vacunal=True).count(),
        'total_urbano': total_urbano,
        'total_rural': total_rural,
        'porcentaje_urbano': porcentaje_urbano,
        'porcentaje_rural': porcentaje_rural,
        'mascotas_por_dia': list(mascotas_por_dia),
    }

    return render(request, 'api/dashboard_usuario.html', context)


@login_required
def dashboard_tecnico(request):
    """Dashboard para técnicos - Solo ven planillas asignadas para revisar"""
    if request.user.tipo_usuario != 'tecnico':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    # Planillas donde el usuario es técnico principal O adicional
    planillas = Planilla.objects.filter(
        Q(tecnico_asignado=request.user) |  # Técnico principal
        Q(tecnicos_adicionales=request.user)  # Técnico adicional
    ).distinct()
    
    responsables = Responsable.objects.filter(
        Q(planilla__tecnico_asignado=request.user) |
        Q(planilla__tecnicos_adicionales=request.user)
    ).distinct().select_related('created_by', 'planilla').order_by('-creado')
    
    mascotas = Mascota.objects.filter(
        Q(responsable__planilla__tecnico_asignado=request.user) |
        Q(responsable__planilla__tecnicos_adicionales=request.user)
    ).distinct()
    
    # Calcular estadísticas de zona
    total_urbano = 0
    total_rural = 0
    
    for responsable in responsables:
        mascotas_responsable = responsable.mascotas.filter(
            Q(responsable__planilla__tecnico_asignado=request.user) |
            Q(responsable__planilla__tecnicos_adicionales=request.user)
        )
        for mascota in mascotas_responsable:
            # Mapear zona del responsable a urbano/rural
            if responsable.zona == 'barrio':
                total_urbano += 1
            elif responsable.zona == 'vereda':
                total_rural += 1
            elif responsable.zona == 'centro poblado':
                # Centro poblado puede ser urbano o rural, por defecto urbano
                total_urbano += 1
            else:
                # Fallback: usar el valor de la planilla si la zona del responsable no está definida
                if responsable.planilla.urbano_rural == 'urbano':
                    total_urbano += 1
                else:
                    total_rural += 1
    
    # Calcular porcentajes
    total_mascotas_count = mascotas.count()
    if total_mascotas_count > 0:
        porcentaje_urbano = round((total_urbano / total_mascotas_count) * 100, 1)
        porcentaje_rural = round((total_rural / total_mascotas_count) * 100, 1)
    else:
        porcentaje_urbano = 0
        porcentaje_rural = 0
    
    # Estadísticas de mascotas por día
    from django.db.models import Count, Min
    from datetime import datetime, timedelta, date
    
    # Obtener la fecha más antigua de mascotas registradas
    primera_mascota = mascotas.aggregate(Min('creado'))['creado__min']
    if primera_mascota:
        fecha_inicio = primera_mascota.date() if hasattr(primera_mascota, 'date') else primera_mascota
    else:
        fecha_inicio = datetime.now().date()
    
    fecha_fin = datetime.now().date()
    
    # Crear lista de todos los días desde el inicio hasta hoy
    dias_completos = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        dias_completos.append(fecha_actual)
        fecha_actual += timedelta(days=1)
    
    # Obtener datos de mascotas por día
    mascotas_query = mascotas.extra(
        select={'dia': 'date(api_mascota.creado)'}
    ).values('dia').annotate(cantidad=Count('id'))
    
    # Crear diccionario con los datos
    mascotas_dict = {item['dia']: item['cantidad'] for item in mascotas_query}
    
    # Crear lista completa con todos los días
    mascotas_por_dia = []
    for dia in dias_completos:
        mascotas_por_dia.append({
            'dia': dia,
            'cantidad': mascotas_dict.get(str(dia), 0)
        })
    
    # Reporte de vacunación por vacunador por día
    vacunadores_en_municipios = Veterinario.objects.filter(
        Q(planillas__in=planillas) |  # Vacunadores principales
        Q(planillas_como_vacunador_adicional__in=planillas),  # Vacunadores adicionales
        tipo_usuario='vacunador'
    ).distinct()
    
    reporte_vacunadores = []
    for vacunador in vacunadores_en_municipios:
        # Contar TODAS las mascotas creadas por este vacunador
        mascotas_vacunador_query = Mascota.objects.filter(
            created_by=vacunador,
            responsable__planilla__in=planillas
        ).extra(select={'dia': 'date(api_mascota.creado)'}).values('dia').annotate(cantidad=Count('id'))
        
        # Crear diccionario con los datos del vacunador
        vacunador_dict = {item['dia']: item['cantidad'] for item in mascotas_vacunador_query}
        
        # Crear lista completa con todos los días para este vacunador
        mascotas_por_dia_vacunador = []
        for dia in dias_completos:
            mascotas_por_dia_vacunador.append({
                'dia': dia,
                'cantidad': vacunador_dict.get(str(dia), 0)
            })
        
        nombre_completo = f"{vacunador.first_name} {vacunador.last_name}".strip()
        
        reporte_vacunadores.append({
            'vacunador': vacunador,
            'nombre_completo': nombre_completo,
            'mascotas_por_dia': mascotas_por_dia_vacunador,
            'total': sum(m['cantidad'] for m in mascotas_por_dia_vacunador)
        })
    
    # Responsables recientes con información del vacunador - mostrar todos, no solo 10
    responsables_recientes = responsables.select_related('created_by', 'planilla')  # Mostrar todos los responsables

    # Obtener municipios únicos para el selector
    municipios_unicos = list(planillas.values_list('municipio', flat=True).distinct().order_by('municipio'))

    context = {
        'user_type': 'Técnico',
        'planillas': planillas,
        'municipios_unicos': municipios_unicos,
        'total_planillas': planillas.count(),
        'total_responsables': responsables.count(),
        'total_mascotas': mascotas.count(),
        'mascotas_con_tarjeta': mascotas.filter(antecedente_vacunal=True).count(),
        'total_urbano': total_urbano,
        'total_rural': total_rural,
        'porcentaje_urbano': porcentaje_urbano,
        'porcentaje_rural': porcentaje_rural,
        'mascotas_por_dia': list(mascotas_por_dia),
        'reporte_vacunadores': reporte_vacunadores,
        'responsables_recientes': responsables_recientes,
        'vacunadores_en_municipios': vacunadores_en_municipios,
    }

    return render(request, 'api/dashboard_usuario.html', context)


@login_required
def dashboard_administrador(request):
    """Dashboard para administradores - Ven todo"""
    if request.user.tipo_usuario != 'administrador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')
    
    # Administradores ven todo
    planillas = Planilla.objects.all()
    responsables = Responsable.objects.all().select_related('created_by', 'planilla').order_by('-creado')
    mascotas = Mascota.objects.all()
    
    # Calcular estadísticas de zona
    total_urbano = 0
    total_rural = 0
    
    for responsable in responsables:
        mascotas_responsable = responsable.mascotas.all()
        for mascota in mascotas_responsable:
            # Mapear zona del responsable a urbano/rural
            if responsable.zona == 'barrio':
                total_urbano += 1
            elif responsable.zona == 'vereda':
                total_rural += 1
            elif responsable.zona == 'centro poblado':
                # Centro poblado puede ser urbano o rural, por defecto urbano
                total_urbano += 1
            else:
                # Fallback: usar el valor de la planilla si la zona del responsable no está definida
                if responsable.planilla.urbano_rural == 'urbano':
                    total_urbano += 1
                else:
                    total_rural += 1
    
    # Calcular porcentajes
    total_mascotas_count = mascotas.count()
    if total_mascotas_count > 0:
        porcentaje_urbano = round((total_urbano / total_mascotas_count) * 100, 1)
        porcentaje_rural = round((total_rural / total_mascotas_count) * 100, 1)
    else:
        porcentaje_urbano = 0
        porcentaje_rural = 0
    
    # Obtener todos los usuarios por tipo con sus relaciones
    vacunadores = Veterinario.objects.filter(tipo_usuario='vacunador').prefetch_related(
        'planillas',  # Planillas como vacunador principal
        'planillas_como_vacunador_adicional',  # Planillas como vacunador adicional
        'responsables_creados',
        'mascotas_creadas'
    ).order_by('username')
    
    tecnicos = Veterinario.objects.filter(tipo_usuario='tecnico').prefetch_related(
        'planillas_asignadas',  # Planillas como técnico principal
        'planillas_como_tecnico_adicional'  # Planillas como técnico adicional
    ).order_by('username')
    
    administradores = Veterinario.objects.filter(tipo_usuario='administrador').order_by('username')
    
    # Responsables recientes con información del vacunador
    responsables_recientes = responsables[:10]  # Últimos 10 responsables

    # Obtener municipios únicos para el selector (administrador ve todos)
    municipios_unicos = list(planillas.values_list('municipio', flat=True).distinct().order_by('municipio'))

    context = {
        'user_type': 'Administrador',
        'planillas': planillas,
        'municipios_unicos': municipios_unicos,
        'total_planillas': planillas.count(),
        'total_responsables': responsables.count(),
        'total_mascotas': mascotas.count(),
        'mascotas_con_tarjeta': mascotas.filter(antecedente_vacunal=True).count(),
        'total_urbano': total_urbano,
        'total_rural': total_rural,
        'porcentaje_urbano': porcentaje_urbano,
        'porcentaje_rural': porcentaje_rural,
        'vacunadores': vacunadores,
        'tecnicos': tecnicos,
        'administradores': administradores,
        'total_vacunadores': vacunadores.count(),
        'total_tecnicos': tecnicos.count(),
        'total_administradores': administradores.count(),
        'responsables_recientes': responsables_recientes,
    }

    return render(request, 'api/dashboard_usuario.html', context)


@login_required
def dashboard_principal(request):
    """Vista principal que redirige según el tipo de usuario"""
    if request.user.tipo_usuario == 'administrador':
        return redirect('dashboard_administrador')
    elif request.user.tipo_usuario == 'vacunador':
        return redirect('dashboard_vacunador')
    elif request.user.tipo_usuario == 'tecnico':
        return redirect('dashboard_tecnico')
    else:
        messages.error(request, 'Tipo de usuario no válido.')
    return redirect('login')


@login_required
def importar_responsables_mascotas(request):
    """
    Vista para importación masiva de responsables y mascotas desde Excel/CSV
    Solo accesible para administradores
    """
    # Verificar que el usuario sea administrador
    if request.user.tipo_usuario != 'administrador':
        messages.error(request, 'No tienes permisos para acceder a esta función.')
        return redirect('dashboard_principal')

    if request.method == 'POST':
        # Verificar que se haya subido un archivo
        if 'archivo' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningún archivo.')
            return redirect('importar_responsables_mascotas')

        archivo = request.FILES['archivo']
        planilla_id = request.POST.get('planilla_id')

        # Verificar que se haya seleccionado una planilla
        if not planilla_id:
            messages.error(request, 'Debes seleccionar una planilla.')
            return redirect('importar_responsables_mascotas')

        try:
            planilla = Planilla.objects.get(id=planilla_id)
        except Planilla.DoesNotExist:
            messages.error(request, 'La planilla seleccionada no existe.')
            return redirect('importar_responsables_mascotas')

        # Verificar extensión del archivo
        extension = archivo.name.split('.')[-1].lower()
        if extension not in ['xlsx', 'xls', 'csv']:
            messages.error(request, 'El archivo debe ser Excel (.xlsx, .xls) o CSV (.csv).')
            return redirect('importar_responsables_mascotas')

        try:
            import openpyxl
            from decimal import Decimal, InvalidOperation

            # Procesar archivo Excel
            if extension in ['xlsx', 'xls']:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                # Leer encabezados (primera fila)
                headers = [cell.value for cell in ws[1]]

                # Diccionario para agrupar mascotas por responsable
                responsables_dict = {}
                errores = []
                linea = 1  # Contador de líneas

                # Procesar cada fila (desde la segunda fila en adelante)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    linea += 1

                    # Saltar filas vacías
                    if not any(row):
                        continue

                    # Crear diccionario con los datos de la fila
                    datos = dict(zip(headers, row))

                    # Validar campos obligatorios del responsable
                    if not datos.get('nombre_responsable'):
                        errores.append(f"Línea {linea}: Falta el nombre del responsable")
                        continue

                    # Crear clave única para el responsable (nombre + teléfono + finca)
                    clave_responsable = (
                        str(datos.get('nombre_responsable', '')).strip(),
                        str(datos.get('telefono', '')).strip(),
                        str(datos.get('finca', '')).strip()
                    )

                    # Si el responsable no existe en el diccionario, agregarlo
                    if clave_responsable not in responsables_dict:
                        responsables_dict[clave_responsable] = {
                            'nombre': datos.get('nombre_responsable', '').strip(),
                            'telefono': datos.get('telefono', '').strip(),
                            'finca': datos.get('finca', '').strip(),
                            'zona': datos.get('zona', 'Sin especificar'),
                            'nombre_zona': datos.get('nombre_zona', 'Sin especificar'),
                            'lote_vacuna': datos.get('lote_vacuna', 'Sin especificar'),
                            'mascotas': []
                        }

                    # Agregar mascota si hay nombre de mascota
                    if datos.get('nombre_mascota'):
                        mascota_data = {
                            'nombre': str(datos.get('nombre_mascota', '')).strip(),
                            'tipo': str(datos.get('tipo_mascota', 'perro')).lower().strip(),
                            'raza': str(datos.get('raza_mascota', 'M')).strip(),
                            'color': str(datos.get('color_mascota', 'Sin especificar')).strip(),
                            'antecedente_vacunal': str(datos.get('antecedente_vacunal', 'NO')).upper().strip() in ['SI', 'SÍ', 'S', 'TRUE', '1'],
                            'esterilizado': str(datos.get('esterilizado', 'NO')).upper().strip() in ['SI', 'SÍ', 'S', 'TRUE', '1'],
                            'latitud': None,
                            'longitud': None
                        }

                        # Procesar coordenadas si existen
                        try:
                            if datos.get('latitud'):
                                mascota_data['latitud'] = Decimal(str(datos.get('latitud')))
                        except (InvalidOperation, ValueError):
                            errores.append(f"Línea {linea}: Latitud inválida")

                        try:
                            if datos.get('longitud'):
                                mascota_data['longitud'] = Decimal(str(datos.get('longitud')))
                        except (InvalidOperation, ValueError):
                            errores.append(f"Línea {linea}: Longitud inválida")

                        responsables_dict[clave_responsable]['mascotas'].append(mascota_data)

                # Crear responsables y mascotas en la base de datos
                responsables_creados = 0
                mascotas_creadas = 0

                for clave, datos_responsable in responsables_dict.items():
                    try:
                        # Crear responsable
                        responsable = Responsable.objects.create(
                            nombre=datos_responsable['nombre'],
                            telefono=datos_responsable['telefono'],
                            finca=datos_responsable['finca'],
                            zona=datos_responsable['zona'],
                            nombre_zona=datos_responsable['nombre_zona'],
                            lote_vacuna=datos_responsable['lote_vacuna'],
                            planilla=planilla,
                            created_by=request.user
                        )
                        responsables_creados += 1

                        # Crear mascotas
                        for mascota_data in datos_responsable['mascotas']:
                            Mascota.objects.create(
                                nombre=mascota_data['nombre'],
                                tipo=mascota_data['tipo'],
                                raza=mascota_data['raza'],
                                color=mascota_data['color'],
                                antecedente_vacunal=mascota_data['antecedente_vacunal'],
                                esterilizado=mascota_data['esterilizado'],
                                latitud=mascota_data['latitud'],
                                longitud=mascota_data['longitud'],
                                responsable=responsable,
                                created_by=request.user
                            )
                            mascotas_creadas += 1

                    except Exception as e:
                        errores.append(f"Error al crear responsable {datos_responsable['nombre']}: {str(e)}")

                # Mostrar resultados
                if errores:
                    messages.warning(request, f"Importación completada con errores. Responsables creados: {responsables_creados}, Mascotas creadas: {mascotas_creadas}. Errores: {'; '.join(errores[:5])}")
                else:
                    messages.success(request, f"Importación exitosa. Se crearon {responsables_creados} responsables y {mascotas_creadas} mascotas.")

                return redirect('dashboard_administrador')

            elif extension == 'csv':
                import csv
                import io

                # Leer archivo CSV
                archivo_texto = archivo.read().decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(archivo_texto))

                # Diccionario para agrupar mascotas por responsable
                responsables_dict = {}
                errores = []
                linea = 1

                for row in csv_reader:
                    linea += 1

                    # Validar campos obligatorios
                    if not row.get('nombre_responsable'):
                        errores.append(f"Línea {linea}: Falta el nombre del responsable")
                        continue

                    # Crear clave única para el responsable
                    clave_responsable = (
                        row.get('nombre_responsable', '').strip(),
                        row.get('telefono', '').strip(),
                        row.get('finca', '').strip()
                    )

                    # Si el responsable no existe, agregarlo
                    if clave_responsable not in responsables_dict:
                        responsables_dict[clave_responsable] = {
                            'nombre': row.get('nombre_responsable', '').strip(),
                            'telefono': row.get('telefono', '').strip(),
                            'finca': row.get('finca', '').strip(),
                            'zona': row.get('zona', 'Sin especificar'),
                            'nombre_zona': row.get('nombre_zona', 'Sin especificar'),
                            'lote_vacuna': row.get('lote_vacuna', 'Sin especificar'),
                            'mascotas': []
                        }

                    # Agregar mascota
                    if row.get('nombre_mascota'):
                        mascota_data = {
                            'nombre': row.get('nombre_mascota', '').strip(),
                            'tipo': row.get('tipo_mascota', 'perro').lower().strip(),
                            'raza': row.get('raza_mascota', 'M').strip(),
                            'color': row.get('color_mascota', 'Sin especificar').strip(),
                            'antecedente_vacunal': row.get('antecedente_vacunal', 'NO').upper().strip() in ['SI', 'SÍ', 'S', 'TRUE', '1'],
                            'esterilizado': row.get('esterilizado', 'NO').upper().strip() in ['SI', 'SÍ', 'S', 'TRUE', '1'],
                            'latitud': None,
                            'longitud': None
                        }

                        # Procesar coordenadas
                        try:
                            if row.get('latitud'):
                                mascota_data['latitud'] = Decimal(row.get('latitud'))
                        except (InvalidOperation, ValueError):
                            errores.append(f"Línea {linea}: Latitud inválida")

                        try:
                            if row.get('longitud'):
                                mascota_data['longitud'] = Decimal(row.get('longitud'))
                        except (InvalidOperation, ValueError):
                            errores.append(f"Línea {linea}: Longitud inválida")

                        responsables_dict[clave_responsable]['mascotas'].append(mascota_data)

                # Crear en la base de datos
                responsables_creados = 0
                mascotas_creadas = 0

                for clave, datos_responsable in responsables_dict.items():
                    try:
                        responsable = Responsable.objects.create(
                            nombre=datos_responsable['nombre'],
                            telefono=datos_responsable['telefono'],
                            finca=datos_responsable['finca'],
                            zona=datos_responsable['zona'],
                            nombre_zona=datos_responsable['nombre_zona'],
                            lote_vacuna=datos_responsable['lote_vacuna'],
                            planilla=planilla,
                            created_by=request.user
                        )
                        responsables_creados += 1

                        for mascota_data in datos_responsable['mascotas']:
                            Mascota.objects.create(
                                nombre=mascota_data['nombre'],
                                tipo=mascota_data['tipo'],
                                raza=mascota_data['raza'],
                                color=mascota_data['color'],
                                antecedente_vacunal=mascota_data['antecedente_vacunal'],
                                esterilizado=mascota_data['esterilizado'],
                                latitud=mascota_data['latitud'],
                                longitud=mascota_data['longitud'],
                                responsable=responsable,
                                created_by=request.user
                            )
                            mascotas_creadas += 1

                    except Exception as e:
                        errores.append(f"Error al crear responsable {datos_responsable['nombre']}: {str(e)}")

                if errores:
                    messages.warning(request, f"Importación completada con errores. Responsables: {responsables_creados}, Mascotas: {mascotas_creadas}. Errores: {'; '.join(errores[:5])}")
                else:
                    messages.success(request, f"Importación exitosa. {responsables_creados} responsables y {mascotas_creadas} mascotas creadas.")

                return redirect('dashboard_administrador')

        except ImportError:
            messages.error(request, 'Error: La librería openpyxl no está instalada. Ejecuta: pip install openpyxl')
            return redirect('importar_responsables_mascotas')

        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('importar_responsables_mascotas')

    # GET request - mostrar formulario
    planillas = Planilla.objects.all().order_by('municipio', 'nombre')
    return render(request, 'api/importar_responsables.html', {'planillas': planillas})


@login_required
def mapa_mascotas(request):
    """Vista para mostrar el mapa con las mascotas georreferenciadas"""
    return render(request, 'api/mapa_mascotas.html')


@api_view(['GET'])
@permission_classes([AllowAny])  # Cambiado temporalmente para permitir acceso
def api_mascotas_georef(request):
    """API endpoint para obtener mascotas con georreferenciación"""
    user = request.user
    
    # Filtrar mascotas según el rol del usuario
    if not user.is_authenticated:
        # Si no está autenticado, mostrar todas las mascotas con georreferenciación
        mascotas = Mascota.objects.filter(
            latitud__isnull=False,
            longitud__isnull=False
        ).select_related('responsable', 'responsable__planilla', 'created_by')
    elif user.tipo_usuario == 'administrador':
        # Administradores ven todas las mascotas con georreferenciación
        mascotas = Mascota.objects.filter(
            latitud__isnull=False,
            longitud__isnull=False
        ).select_related('responsable', 'responsable__planilla', 'created_by')
    elif user.tipo_usuario == 'tecnico':
        # Técnicos ven las mascotas de sus municipios asignados
        mascotas = Mascota.objects.filter(
            Q(responsable__planilla__tecnico_asignado=user) |
            Q(responsable__planilla__tecnicos_adicionales=user),
            latitud__isnull=False,
            longitud__isnull=False
        ).select_related('responsable', 'responsable__planilla', 'created_by').distinct()
    elif user.tipo_usuario == 'vacunador':
        # Vacunadores ven solo las mascotas que ellos crearon
        mascotas = Mascota.objects.filter(
            created_by=user,
            latitud__isnull=False,
            longitud__isnull=False
        ).select_related('responsable', 'responsable__planilla')
    else:
        mascotas = Mascota.objects.none()
    
    # Preparar datos para el mapa
    mascotas_data = []
    for mascota in mascotas:
        mascotas_data.append({
            'id': mascota.id,
            'nombre': mascota.nombre,
            'tipo': mascota.get_tipo_display(),
            'raza': mascota.raza,
            'color': mascota.color,
            'esterilizado': mascota.esterilizado,
            'antecedente_vacunal': mascota.antecedente_vacunal,
            'latitud': float(mascota.latitud),
            'longitud': float(mascota.longitud),
            'responsable': {
                'nombre': mascota.responsable.nombre,
                'telefono': mascota.responsable.telefono,
                'finca': mascota.responsable.finca,
                'zona': mascota.responsable.zona,
                'nombre_zona': mascota.responsable.nombre_zona,
            },
            'municipio': mascota.responsable.planilla.municipio,
            'zona_tipo': mascota.responsable.planilla.urbano_rural,
            'vacunador': mascota.created_by.username if mascota.created_by else 'N/A',
            'fecha_registro': mascota.creado.strftime('%Y-%m-%d %H:%M'),
            'foto_url': mascota.foto.url if mascota.foto else None
        })
    
    return Response({
        'mascotas': mascotas_data,
        'total': len(mascotas_data)
    })


# ========== VISTAS PARA REGISTRO DE PÉRDIDAS ==========

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def registro_perdidas_list(request):
    """
    GET: Lista todos los registros de pérdidas del usuario actual
    POST: Crea un nuevo registro de pérdida
    """
    if request.method == 'GET':
        # Filtrar por usuario actual
        perdidas = RegistroPerdidas.objects.filter(registrado_por=request.user).order_by('-fecha_registro')
        data = []
        for perdida in perdidas:
            data.append({
                'id': perdida.id,
                'cantidad': perdida.cantidad,
                'lote_vacuna': perdida.lote_vacuna,
                'motivo': perdida.motivo,
                'fecha_perdida': perdida.fecha_perdida.strftime('%Y-%m-%d'),
                'fecha_registro': perdida.fecha_registro.strftime('%Y-%m-%d %H:%M'),
                'latitud': float(perdida.latitud) if perdida.latitud else None,
                'longitud': float(perdida.longitud) if perdida.longitud else None,
                'foto_url': perdida.foto.url if perdida.foto else None,
                'sincronizado': perdida.sincronizado,
                'uuid_local': perdida.uuid_local
            })
        return Response(data)
    
    elif request.method == 'POST':
        try:
            # Crear nuevo registro de pérdida
            data = request.data
            
            perdida = RegistroPerdidas(
                registrado_por=request.user,
                cantidad=data.get('cantidad'),
                lote_vacuna=data.get('lote_vacuna'),
                motivo=data.get('motivo', ''),
                latitud=data.get('latitud'),
                longitud=data.get('longitud'),
                uuid_local=data.get('uuid_local')
            )
            
            # Manejar foto si viene en base64
            foto_base64 = data.get('foto_base64')
            if foto_base64:
                try:
                    # Si viene con prefijo data:image, procesarlo
                    if isinstance(foto_base64, str) and foto_base64.startswith('data:image'):
                        # Remover el prefijo data:image/jpeg;base64,
                        format, imgstr = foto_base64.split(';base64,')
                        ext = format.split('/')[-1]
                    else:
                        # Si es solo base64 sin prefijo
                        imgstr = foto_base64
                        ext = 'jpg'  # Asumimos jpg por defecto
                    
                    # Decodificar y guardar la imagen
                    foto = ContentFile(base64.b64decode(imgstr), name=f'perdida_{perdida.uuid_local or datetime.now().timestamp()}.{ext}')
                    perdida.foto = foto
                except Exception as e:
                    print(f"Error procesando imagen: {e}")
            
            perdida.save()
            
            return Response({
                'id': perdida.id,
                'uuid_local': perdida.uuid_local,
                'message': 'Registro de pérdida creado exitosamente'
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_perdidas(request):
    """
    Obtiene estadísticas de pérdidas por usuario/municipio
    """
    user = request.user

    # Si es administrador, ve todo
    if user.tipo_usuario == 'administrador':
        perdidas = RegistroPerdidas.objects.all()
    # Si es técnico, ve las de su municipio
    elif user.tipo_usuario == 'tecnico':
        # Obtener planillas del técnico
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).distinct()
        # Obtener usuarios de esas planillas
        usuarios = set()
        for p in planillas:
            usuarios.update(p.get_all_vacunadores())
            usuarios.update(p.get_all_tecnicos())
        perdidas = RegistroPerdidas.objects.filter(registrado_por__in=usuarios)
    else:
        # Vacunador solo ve las suyas
        perdidas = RegistroPerdidas.objects.filter(registrado_por=user)

    # Calcular estadísticas
    total_perdidas = perdidas.count()
    total_vacunas_perdidas = sum(p.cantidad for p in perdidas)

    # Agrupar por lote
    perdidas_por_lote = {}
    for p in perdidas:
        if p.lote_vacuna not in perdidas_por_lote:
            perdidas_por_lote[p.lote_vacuna] = 0
        perdidas_por_lote[p.lote_vacuna] += p.cantidad

    return Response({
        'total_registros': total_perdidas,
        'total_vacunas_perdidas': total_vacunas_perdidas,
        'perdidas_por_lote': perdidas_por_lote,
        'tipo_usuario': user.tipo_usuario
    })


# ========== REPORTE PDF PARA VACUNADORES ==========

@login_required
def reporte_vacunador_pdf(request):
    """Generar reporte PDF de vacunaciones diarias para vacunadores"""
    if request.user.tipo_usuario != 'vacunador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from io import BytesIO
    from datetime import datetime, date
    from django.db.models import Count

    # Obtener fecha específica del parámetro GET o usar fecha actual
    fecha_str = request.GET.get('fecha')
    if fecha_str:
        try:
            fecha_reporte = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_reporte = date.today()
    else:
        fecha_reporte = date.today()

    # Obtener mascotas del vacunador para la fecha específica
    mascotas = Mascota.objects.filter(
        created_by=request.user,
        creado__date=fecha_reporte
    ).select_related('responsable', 'responsable__planilla')

    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)

    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()

    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.darkblue,
        spaceAfter=20,
        alignment=1  # Centrado
    )

    # Título del reporte
    title = Paragraph(f"Reporte de Vacunación Diaria<br/>Vacunador: {request.user.username}<br/>Fecha: {fecha_reporte.strftime('%d/%m/%Y')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))

    # Resumen estadístico
    total_mascotas = mascotas.count()
    total_perros = mascotas.filter(tipo='perro').count()
    total_gatos = mascotas.filter(tipo='gato').count()
    con_tarjeta = mascotas.filter(antecedente_vacunal=True).count()
    sin_tarjeta = mascotas.filter(antecedente_vacunal=False).count()

    resumen_data = [
        ['Resumen del Día', ''],
        ['Total de Mascotas Vacunadas:', str(total_mascotas)],
        ['Perros:', str(total_perros)],
        ['Gatos:', str(total_gatos)],
        ['Con Tarjeta Previa:', str(con_tarjeta)],
        ['Sin Tarjeta Previa:', str(sin_tarjeta)],
    ]

    resumen_table = Table(resumen_data, colWidths=[3*inch, 1*inch])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 30))

    if total_mascotas > 0:
        # Tabla detallada de mascotas
        data = [['#', 'Mascota', 'Tipo', 'Responsable', 'Municipio', 'Zona', 'Tarjeta Previa']]

        for i, mascota in enumerate(mascotas, 1):
            data.append([
                str(i),
                mascota.nombre,
                mascota.get_tipo_display(),
                mascota.responsable.nombre,
                mascota.responsable.planilla.municipio,
                mascota.responsable.zona.title(),
                'Sí' if mascota.antecedente_vacunal else 'No'
            ])

        table = Table(data, colWidths=[0.5*inch, 1.2*inch, 0.8*inch, 1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(Paragraph("Detalle de Mascotas Vacunadas:", styles['Heading2']))
        elements.append(Spacer(1, 10))
        elements.append(table)
    else:
        elements.append(Paragraph("No se registraron vacunaciones en esta fecha.", styles['Normal']))

    # Pie de página
    elements.append(Spacer(1, 30))
    footer_text = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - Sistema de Vacunación VetControl"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)

    # Construir el PDF
    doc.build(elements)

    # Obtener el contenido del buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_vacunacion_{request.user.username}_{fecha_reporte.strftime("%Y-%m-%d")}.pdf"'
    response.write(pdf)

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def arbol_reportes(request):
    """
    API para obtener datos del árbol jerárquico de reportes
    Meta → Municipio → Vacunador → Día → Responsable → Mascota
    Solo para administradores

    Parámetros de filtrado:
    - fecha_inicio: YYYY-MM-DD
    - fecha_fin: YYYY-MM-DD
    - municipio: nombre del municipio
    - vacunador: ID del vacunador
    - tipo_mascota: 'perro' o 'gato'
    - solo_con_antecedente: true/false
    - solo_esterilizados: true/false
    """
    if request.user.tipo_usuario != 'administrador':
        return Response({'error': 'Solo administradores pueden acceder a esta vista'}, status=status.HTTP_403_FORBIDDEN)

    from django.db.models import Count, Q
    from collections import defaultdict
    import json
    from datetime import datetime

    # Obtener parámetros de filtrado
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    municipio_filtro = request.GET.get('municipio')
    vacunador_filtro = request.GET.get('vacunador')
    tipo_mascota_filtro = request.GET.get('tipo_mascota')
    solo_con_antecedente = request.GET.get('solo_con_antecedente') == 'true'
    solo_esterilizados = request.GET.get('solo_esterilizados') == 'true'

    # Estructura del árbol: Meta -> Municipio -> Vacunador -> Día -> Responsable -> Mascota
    arbol = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))

    # Construir queryset base
    mascotas_query = Mascota.objects.select_related(
        'responsable',
        'responsable__planilla',
        'created_by'
    )

    # Aplicar filtros
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            mascotas_query = mascotas_query.filter(creado__date__gte=fecha_inicio_obj)
        except ValueError:
            pass

    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            mascotas_query = mascotas_query.filter(creado__date__lte=fecha_fin_obj)
        except ValueError:
            pass

    if municipio_filtro:
        mascotas_query = mascotas_query.filter(responsable__planilla__municipio__icontains=municipio_filtro)

    if vacunador_filtro:
        try:
            vacunador_id = int(vacunador_filtro)
            mascotas_query = mascotas_query.filter(created_by_id=vacunador_id)
        except (ValueError, TypeError):
            pass

    if tipo_mascota_filtro and tipo_mascota_filtro in ['perro', 'gato']:
        mascotas_query = mascotas_query.filter(tipo=tipo_mascota_filtro)

    if solo_con_antecedente:
        mascotas_query = mascotas_query.filter(antecedente_vacunal=True)

    if solo_esterilizados:
        mascotas_query = mascotas_query.filter(esterilizado=True)

    # Ordenar los resultados
    mascotas = mascotas_query.order_by(
        'responsable__planilla__municipio',
        'created_by__username',
        'creado__date',
        'responsable__nombre',
        'nombre'
    )

    # Construir el árbol
    for mascota in mascotas:
        municipio = mascota.responsable.planilla.municipio
        vacunador = mascota.created_by.username if mascota.created_by else 'Sin asignar'
        fecha = mascota.creado.strftime('%Y-%m-%d')
        responsable_nombre = mascota.responsable.nombre

        # Datos de la mascota
        mascota_data = {
            'id': mascota.id,
            'nombre': mascota.nombre,
            'tipo': mascota.tipo,
            'raza': mascota.raza,
            'color': mascota.color,
            'antecedente_vacunal': mascota.antecedente_vacunal,
            'esterilizado': mascota.esterilizado,
            'creado': mascota.creado.isoformat(),
            'responsable': {
                'id': mascota.responsable.id,
                'nombre': mascota.responsable.nombre,
                'telefono': mascota.responsable.telefono,
                'finca': mascota.responsable.finca,
                'zona': mascota.responsable.zona,
                'nombre_zona': mascota.responsable.nombre_zona,
                'lote_vacuna': mascota.responsable.lote_vacuna,
            }
        }

        arbol[municipio][vacunador][fecha][responsable_nombre].append(mascota_data)

    # Convertir a estructura de árbol con conteos
    resultado = []
    total_mascotas = 0

    for municipio, vacunadores in arbol.items():
        municipio_mascotas = 0
        vacunadores_list = []

        for vacunador, fechas in vacunadores.items():
            vacunador_mascotas = 0
            fechas_list = []

            for fecha, responsables in fechas.items():
                fecha_mascotas = 0
                responsables_list = []

                for responsable_nombre, mascotas_list in responsables.items():
                    responsable_mascotas = len(mascotas_list)
                    fecha_mascotas += responsable_mascotas

                    # Contar perros y gatos para este responsable
                    responsable_perros = len([m for m in mascotas_list if m['tipo'] == 'perro'])
                    responsable_gatos = len([m for m in mascotas_list if m['tipo'] == 'gato'])

                    responsables_list.append({
                        'id': mascotas_list[0]['responsable']['id'] if mascotas_list else None,
                        'nombre': responsable_nombre,
                        'telefono': mascotas_list[0]['responsable']['telefono'] if mascotas_list else '',
                        'finca': mascotas_list[0]['responsable']['finca'] if mascotas_list else '',
                        'zona': mascotas_list[0]['responsable']['zona'] if mascotas_list else '',
                        'total_mascotas': responsable_mascotas,
                        'total_perros': responsable_perros,
                        'total_gatos': responsable_gatos,
                        'mascotas': mascotas_list
                    })

                vacunador_mascotas += fecha_mascotas

                # Contar perros y gatos para esta fecha
                fecha_perros = sum(r['total_perros'] for r in responsables_list)
                fecha_gatos = sum(r['total_gatos'] for r in responsables_list)

                fechas_list.append({
                    'fecha': fecha,
                    'total_mascotas': fecha_mascotas,
                    'total_perros': fecha_perros,
                    'total_gatos': fecha_gatos,
                    'responsables': responsables_list
                })

            municipio_mascotas += vacunador_mascotas

            # Contar perros y gatos para este vacunador
            vacunador_perros = sum(f['total_perros'] for f in fechas_list)
            vacunador_gatos = sum(f['total_gatos'] for f in fechas_list)

            vacunadores_list.append({
                'nombre': vacunador,
                'total_mascotas': vacunador_mascotas,
                'total_perros': vacunador_perros,
                'total_gatos': vacunador_gatos,
                'fechas': fechas_list
            })

        total_mascotas += municipio_mascotas

        # Contar perros y gatos para este municipio
        municipio_perros = sum(v['total_perros'] for v in vacunadores_list)
        municipio_gatos = sum(v['total_gatos'] for v in vacunadores_list)

        resultado.append({
            'municipio': municipio,
            'total_mascotas': municipio_mascotas,
            'total_perros': municipio_perros,
            'total_gatos': municipio_gatos,
            'vacunadores': vacunadores_list
        })

    # Ordenar por municipio
    resultado.sort(key=lambda x: x['municipio'])

    # Obtener información adicional para filtros
    filtros_aplicados = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'municipio': municipio_filtro,
        'vacunador': vacunador_filtro,
        'tipo_mascota': tipo_mascota_filtro,
        'solo_con_antecedente': solo_con_antecedente,
        'solo_esterilizados': solo_esterilizados,
    }

    return Response({
        'total_mascotas': total_mascotas,
        'total_municipios': len(resultado),
        'filtros_aplicados': filtros_aplicados,
        'arbol': resultado
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def estadisticas_generales(request):
    """
    API para obtener estadísticas generales del sistema
    Solo para administradores
    """
    if request.user.tipo_usuario != 'administrador':
        return Response({'error': 'Solo administradores pueden acceder a esta vista'}, status=status.HTTP_403_FORBIDDEN)

    from django.db.models import Count, Q

    # Estadísticas generales
    total_mascotas = Mascota.objects.count()
    total_responsables = Responsable.objects.count()
    total_planillas = Planilla.objects.count()
    total_municipios = Planilla.objects.values('municipio').distinct().count()

    # Estadísticas por tipo
    mascotas_por_tipo = Mascota.objects.values('tipo').annotate(count=Count('id'))
    mascotas_con_antecedente = Mascota.objects.filter(antecedente_vacunal=True).count()
    mascotas_esterilizadas = Mascota.objects.filter(esterilizado=True).count()

    # Estadísticas por usuario
    usuarios_activos = Veterinario.objects.filter(
        Q(mascotas_creadas__isnull=False) | Q(responsables_creados__isnull=False)
    ).distinct().count()

    return Response({
        'totales': {
            'mascotas': total_mascotas,
            'responsables': total_responsables,
            'planillas': total_planillas,
            'municipios': total_municipios,
            'usuarios_activos': usuarios_activos
        },
        'mascotas_por_tipo': list(mascotas_por_tipo),
        'mascotas_con_antecedente': mascotas_con_antecedente,
        'mascotas_esterilizadas': mascotas_esterilizadas,
        'porcentajes': {
            'con_antecedente': round((mascotas_con_antecedente / total_mascotas * 100), 2) if total_mascotas > 0 else 0,
            'esterilizadas': round((mascotas_esterilizadas / total_mascotas * 100), 2) if total_mascotas > 0 else 0,
        }
    })


@login_required
def arbol_reportes_view(request):
    """
    Vista web para mostrar el árbol jerárquico de reportes
    Solo para administradores
    """
    if request.user.tipo_usuario != 'administrador':
        messages.error(request, 'Solo administradores pueden acceder al árbol de reportes.')
        return redirect('login')

    return render(request, 'api/arbol_reportes.html')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def opciones_filtros_reportes(request):
    """
    API para obtener las opciones disponibles para los filtros del árbol de reportes
    Solo para administradores
    """
    if request.user.tipo_usuario != 'administrador':
        return Response({'error': 'Solo administradores pueden acceder a esta vista'}, status=status.HTTP_403_FORBIDDEN)

    from django.db.models import Min, Max

    # Obtener municipios únicos
    municipios = list(Planilla.objects.values_list('municipio', flat=True).distinct().order_by('municipio'))

    # Obtener vacunadores únicos (que hayan creado mascotas)
    vacunadores = list(Veterinario.objects.filter(
        mascotas_creadas__isnull=False
    ).distinct().values('id', 'username', 'first_name', 'last_name').order_by('username'))

    # Formatear nombres de vacunadores
    for vacunador in vacunadores:
        nombre_completo = f"{vacunador['first_name']} {vacunador['last_name']}".strip()
        vacunador['nombre_display'] = nombre_completo if nombre_completo else vacunador['username']

    # Obtener rango de fechas disponibles
    fechas_range = Mascota.objects.aggregate(
        fecha_min=Min('creado__date'),
        fecha_max=Max('creado__date')
    )

    return Response({
        'municipios': municipios,
        'vacunadores': vacunadores,
        'tipos_mascota': [
            {'value': 'perro', 'label': '🐕 Perros'},
            {'value': 'gato', 'label': '🐱 Gatos'}
        ],
        'fecha_min': fechas_range['fecha_min'],
        'fecha_max': fechas_range['fecha_max'],
        'filtros_especiales': [
            {'key': 'solo_con_antecedente', 'label': 'Solo con antecedente vacunal'},
            {'key': 'solo_esterilizados', 'label': 'Solo esterilizados'}
        ]
    })


# API endpoints para edición de responsables y mascotas
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_responsable(request, responsable_id):
    """
    Actualizar datos de un responsable
    """
    try:
        responsable = get_object_or_404(Responsable, id=responsable_id)

        # Verificar permisos
        if request.user.tipo_usuario not in ['administrador', 'tecnico']:
            return Response(
                {'error': 'No tiene permisos para editar responsables'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Actualizar campos
        responsable.nombre = request.data.get('nombre', responsable.nombre)
        responsable.telefono = request.data.get('telefono', responsable.telefono)
        responsable.finca = request.data.get('finca', responsable.finca)
        responsable.zona = request.data.get('zona', responsable.zona)
        responsable.nombre_zona = request.data.get('nombre_zona', responsable.nombre_zona)
        responsable.lote_vacuna = request.data.get('lote_vacuna', responsable.lote_vacuna)

        responsable.save()

        return Response({
            'message': 'Responsable actualizado correctamente',
            'responsable': {
                'id': responsable.id,
                'nombre': responsable.nombre,
                'telefono': responsable.telefono,
                'finca': responsable.finca,
                'zona': responsable.zona,
                'nombre_zona': responsable.nombre_zona,
                'lote_vacuna': responsable.lote_vacuna
            }
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_mascota(request, mascota_id):
    """
    Eliminar una mascota
    """
    try:
        mascota = get_object_or_404(Mascota, id=mascota_id)

        # Verificar permisos
        print(f"DEBUG DELETE MASCOTA - Usuario: {request.user.username}, Tipo: {getattr(request.user, 'tipo_usuario', 'NO_TIENE_TIPO')}, Autenticado: {request.user.is_authenticated}")

        if not request.user.is_authenticated:
            return Response(
                {'error': 'Usuario no autenticado'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if not hasattr(request.user, 'tipo_usuario'):
            return Response(
                {'error': 'Usuario sin tipo definido'},
                status=status.HTTP_403_FORBIDDEN
            )

        if request.user.tipo_usuario not in ['administrador', 'tecnico']:
            return Response(
                {'error': f'No tiene permisos para eliminar mascotas. Su tipo de usuario es: {request.user.tipo_usuario}'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Guardar información antes de eliminar
        mascota_info = {
            'id': mascota.id,
            'nombre': mascota.nombre,
            'tipo': mascota.tipo,
            'responsable': mascota.responsable.nombre if mascota.responsable else 'N/A'
        }

        mascota.delete()

        return Response({
            'message': f'Mascota "{mascota_info["nombre"]}" eliminada correctamente',
            'mascota_eliminada': mascota_info
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_responsable(request, responsable_id):
    """
    Eliminar un responsable y todas sus mascotas
    """
    try:
        responsable = get_object_or_404(Responsable, id=responsable_id)

        # Verificar permisos
        print(f"DEBUG DELETE RESPONSABLE - Usuario: {request.user.username}, Tipo: {getattr(request.user, 'tipo_usuario', 'NO_TIENE_TIPO')}, Autenticado: {request.user.is_authenticated}")

        if not request.user.is_authenticated:
            return Response(
                {'error': 'Usuario no autenticado'},
                status=status.HTTP_401_UNAUTHORIZED
            )

        if not hasattr(request.user, 'tipo_usuario'):
            return Response(
                {'error': 'Usuario sin tipo definido'},
                status=status.HTTP_403_FORBIDDEN
            )

        if request.user.tipo_usuario not in ['administrador', 'tecnico']:
            return Response(
                {'error': f'No tiene permisos para eliminar responsables. Su tipo de usuario es: {request.user.tipo_usuario}'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Contar mascotas antes de eliminar
        total_mascotas = responsable.mascotas.count()

        # Guardar información antes de eliminar
        responsable_info = {
            'id': responsable.id,
            'nombre': responsable.nombre,
            'telefono': responsable.telefono,
            'finca': responsable.finca,
            'total_mascotas': total_mascotas
        }

        responsable.delete()  # Esto también eliminará las mascotas por CASCADE

        return Response({
            'message': f'Responsable "{responsable_info["nombre"]}" y {total_mascotas} mascota(s) eliminado(s) correctamente',
            'responsable_eliminado': responsable_info
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_fecha_creacion_mascota(request, mascota_id):
    """
    Actualizar fecha de creación de una mascota
    """
    try:
        mascota = get_object_or_404(Mascota, id=mascota_id)

        # Verificar permisos
        if request.user.tipo_usuario not in ['administrador', 'tecnico']:
            return Response(
                {'error': 'No tiene permisos para editar fechas de mascotas'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Obtener nueva fecha
        nueva_fecha = request.data.get('fecha_creacion')
        if not nueva_fecha:
            return Response(
                {'error': 'Se requiere el campo fecha_creacion'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.utils import timezone
        from datetime import datetime

        # Convertir string a datetime
        try:
            # Intentar varios formatos
            try:
                fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d')
                except ValueError:
                    fecha_obj = datetime.strptime(nueva_fecha, '%d/%m/%Y')

            # Hacer timezone aware
            fecha_obj = timezone.make_aware(fecha_obj)

        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD o YYYY-MM-DD HH:MM:SS'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Actualizar fecha
        mascota.creado = fecha_obj
        mascota.save()

        return Response({
            'message': 'Fecha de creación actualizada correctamente',
            'mascota': {
                'id': mascota.id,
                'nombre': mascota.nombre,
                'fecha_creacion_anterior': request.data.get('fecha_anterior', 'N/A'),
                'fecha_creacion_nueva': mascota.creado.isoformat()
            }
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_fecha_creacion_responsable(request, responsable_id):
    """
    Actualizar fecha de creación de un responsable
    """
    try:
        responsable = get_object_or_404(Responsable, id=responsable_id)

        # Verificar permisos
        if request.user.tipo_usuario not in ['administrador', 'tecnico']:
            return Response(
                {'error': 'No tiene permisos para editar fechas de responsables'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Obtener nueva fecha
        nueva_fecha = request.data.get('fecha_creacion')
        if not nueva_fecha:
            return Response(
                {'error': 'Se requiere el campo fecha_creacion'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.utils import timezone
        from datetime import datetime

        # Convertir string a datetime
        try:
            # Intentar varios formatos
            try:
                fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                try:
                    fecha_obj = datetime.strptime(nueva_fecha, '%Y-%m-%d')
                except ValueError:
                    fecha_obj = datetime.strptime(nueva_fecha, '%d/%m/%Y')

            # Hacer timezone aware
            fecha_obj = timezone.make_aware(fecha_obj)

        except ValueError:
            return Response(
                {'error': 'Formato de fecha inválido. Use YYYY-MM-DD o YYYY-MM-DD HH:MM:SS'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Actualizar fecha
        responsable.creado = fecha_obj
        responsable.save()

        return Response({
            'message': 'Fecha de creación actualizada correctamente',
            'responsable': {
                'id': responsable.id,
                'nombre': responsable.nombre,
                'fecha_creacion_anterior': request.data.get('fecha_anterior', 'N/A'),
                'fecha_creacion_nueva': responsable.creado.isoformat()
            }
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_mascota(request, mascota_id):
    """
    Actualizar datos de una mascota
    """
    try:
        mascota = get_object_or_404(Mascota, id=mascota_id)

        # Verificar permisos
        if request.user.tipo_usuario not in ['administrador', 'tecnico', 'vacunador']:
            return Response(
                {'error': 'No tiene permisos para editar mascotas'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Actualizar campos
        mascota.nombre = request.data.get('nombre', mascota.nombre)
        mascota.tipo = request.data.get('tipo', mascota.tipo)
        mascota.raza = request.data.get('raza', mascota.raza)
        mascota.color = request.data.get('color', mascota.color)
        mascota.antecedente_vacunal = request.data.get('antecedente_vacunal', mascota.antecedente_vacunal)
        mascota.esterilizado = request.data.get('esterilizado', mascota.esterilizado)

        mascota.save()

        return Response({
            'message': 'Mascota actualizada correctamente',
            'mascota': {
                'id': mascota.id,
                'nombre': mascota.nombre,
                'tipo': mascota.tipo,
                'raza': mascota.raza,
                'color': mascota.color,
                'antecedente_vacunal': mascota.antecedente_vacunal,
                'esterilizado': mascota.esterilizado
            }
        })

    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )


@login_required
def reporte_estadistico_vacunacion_pdf(request):
    """
    Generar reporte estadístico de vacunación por día y vacunador
    Muestra: Día, Vacunador, Perros Total, Perros Rural, Perros Urbano, Gatos Total, Gatos Rural, Gatos Urbano
    Disponible para técnicos, vacunadores y administradores
    """
    user = request.user

    # Verificar permisos
    if user.tipo_usuario not in ['tecnico', 'vacunador', 'administrador']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Importar fix de compatibilidad para ReportLab con Python 3.8
    from . import pdf_utils

    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from io import BytesIO
    from datetime import datetime, date
    from django.db.models import Q, Count
    from collections import defaultdict

    # Obtener parámetro de municipio
    municipio_filtro = request.GET.get('municipio')

    # Obtener planillas según el rol del usuario
    if user.tipo_usuario == 'administrador':
        # Administradores ven todos los municipios
        planillas = Planilla.objects.all()
    elif user.tipo_usuario == 'tecnico':
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).distinct()
    elif user.tipo_usuario == 'vacunador':
        planillas = Planilla.objects.filter(
            Q(assigned_to=user) |
            Q(vacunadores_adicionales=user)
        ).distinct()
    else:
        planillas = Planilla.objects.none()

    # Filtrar por municipio si se proporciona
    if municipio_filtro:
        planillas = planillas.filter(municipio__icontains=municipio_filtro)

    # Obtener mascotas del municipio
    if user.tipo_usuario in ['administrador', 'tecnico']:
        # Administradores y técnicos ven todas las mascotas de sus municipios
        mascotas = Mascota.objects.filter(
            responsable__planilla__in=planillas
        ).select_related('responsable', 'responsable__planilla', 'created_by')
    else:
        # Vacunadores solo ven las mascotas que crearon
        mascotas = Mascota.objects.filter(
            created_by=user,
            responsable__planilla__in=planillas
        ).select_related('responsable', 'responsable__planilla')

    # Estructura de datos: {fecha: {vacunador: {estadisticas}}}
    datos_reporte = defaultdict(lambda: defaultdict(lambda: {
        'perros_total': 0,
        'perros_rural': 0,
        'perros_urbano': 0,
        'gatos_total': 0,
        'gatos_rural': 0,
        'gatos_urbano': 0,
    }))

    # Procesar cada mascota
    for mascota in mascotas:
        fecha = mascota.creado.date()
        vacunador_nombre = mascota.created_by.username if mascota.created_by else 'Sin asignar'
        responsable = mascota.responsable

        # Determinar si es rural o urbano
        # CP (centro poblado) y V (vereda) = rural
        # B (barrio) = urbano
        es_rural = responsable.zona in ['vereda', 'centro poblado']
        es_urbano = responsable.zona == 'barrio'

        # Si no está definido, usar el valor de la planilla
        if responsable.zona not in ['vereda', 'centro poblado', 'barrio']:
            es_rural = responsable.planilla.urbano_rural == 'rural'
            es_urbano = responsable.planilla.urbano_rural == 'urbano'

        # Contabilizar según tipo de mascota
        if mascota.tipo == 'perro':
            datos_reporte[fecha][vacunador_nombre]['perros_total'] += 1
            if es_rural:
                datos_reporte[fecha][vacunador_nombre]['perros_rural'] += 1
            if es_urbano:
                datos_reporte[fecha][vacunador_nombre]['perros_urbano'] += 1
        elif mascota.tipo == 'gato':
            datos_reporte[fecha][vacunador_nombre]['gatos_total'] += 1
            if es_rural:
                datos_reporte[fecha][vacunador_nombre]['gatos_rural'] += 1
            if es_urbano:
                datos_reporte[fecha][vacunador_nombre]['gatos_urbano'] += 1

    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)

    elements = []
    styles = getSampleStyleSheet()

    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=14,
        textColor=colors.darkblue,
        spaceAfter=12,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=10,
        alignment=1
    )

    # Título del reporte
    municipio_texto = municipio_filtro if municipio_filtro else "Todos los municipios"
    title = Paragraph(
        f"Reporte Estadístico de Vacunación por Día y Vacunador<br/>Municipio: {municipio_texto}",
        title_style
    )
    elements.append(title)

    subtitle = Paragraph(
        f"{user.get_tipo_usuario_display()}: {user.username} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        subtitle_style
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 20))

    if not datos_reporte:
        no_data = Paragraph(
            "No hay datos de vacunación para los criterios seleccionados.",
            styles['Normal']
        )
        elements.append(no_data)
    else:
        # Crear tabla de estadísticas
        data = [[
            'Fecha',
            'Vacunador',
            'Perros\nTotal',
            'Perros\nRural',
            'Perros\nUrbano',
            'Gatos\nTotal',
            'Gatos\nRural',
            'Gatos\nUrbano',
            'Total\nMascotas'
        ]]

        # Totales generales
        totales = {
            'perros_total': 0,
            'perros_rural': 0,
            'perros_urbano': 0,
            'gatos_total': 0,
            'gatos_rural': 0,
            'gatos_urbano': 0,
        }

        # Ordenar por fecha
        for fecha in sorted(datos_reporte.keys()):
            vacunadores_dict = datos_reporte[fecha]

            # Ordenar vacunadores alfabéticamente
            for vacunador in sorted(vacunadores_dict.keys()):
                stats = vacunadores_dict[vacunador]

                total_mascotas = stats['perros_total'] + stats['gatos_total']

                data.append([
                    fecha.strftime('%d/%m/%Y'),
                    vacunador,
                    str(stats['perros_total']),
                    str(stats['perros_rural']),
                    str(stats['perros_urbano']),
                    str(stats['gatos_total']),
                    str(stats['gatos_rural']),
                    str(stats['gatos_urbano']),
                    str(total_mascotas)
                ])

                # Acumular totales
                totales['perros_total'] += stats['perros_total']
                totales['perros_rural'] += stats['perros_rural']
                totales['perros_urbano'] += stats['perros_urbano']
                totales['gatos_total'] += stats['gatos_total']
                totales['gatos_rural'] += stats['gatos_rural']
                totales['gatos_urbano'] += stats['gatos_urbano']

        # Agregar fila de totales
        total_general = totales['perros_total'] + totales['gatos_total']
        data.append([
            'TOTALES',
            '',
            str(totales['perros_total']),
            str(totales['perros_rural']),
            str(totales['perros_urbano']),
            str(totales['gatos_total']),
            str(totales['gatos_rural']),
            str(totales['gatos_urbano']),
            str(total_general)
        ])

        # Configurar tabla
        table = Table(data, colWidths=[
            0.9*inch,   # Fecha
            1.2*inch,   # Vacunador
            0.7*inch,   # Perros Total
            0.7*inch,   # Perros Rural
            0.7*inch,   # Perros Urbano
            0.7*inch,   # Gatos Total
            0.7*inch,   # Gatos Rural
            0.7*inch,   # Gatos Urbano
            0.8*inch    # Total Mascotas
        ])

        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Contenido
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f7f7f7')]),

            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 8),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1976d2')),
        ]))

        elements.append(table)

    # Leyenda
    elements.append(Spacer(1, 20))
    leyenda_text = "<i><b>Nota:</b> Rural incluye Veredas y Centros Poblados. Urbano incluye Barrios.</i>"
    leyenda = Paragraph(leyenda_text, subtitle_style)
    elements.append(leyenda)

    # Pie de página
    footer_text = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - Sistema VetControl"
    footer = Paragraph(footer_text, subtitle_style)
    elements.append(footer)

    # Construir el PDF
    doc.build(elements)

    # Obtener el contenido del buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'reporte_estadistico_{municipio_filtro or "todos"}_{user.username}_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)

    return response


@login_required
def imprimir_planilla_municipio_pdf(request):
    """
    Generar PDF de planillas por municipio con máximo 20 mascotas por hoja
    Disponible para técnicos, vacunadores y administradores
    """
    user = request.user

    # Verificar permisos
    if user.tipo_usuario not in ['tecnico', 'vacunador', 'administrador']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Importar fix de compatibilidad para ReportLab con Python 3.8
    from . import pdf_utils

    from django.http import HttpResponse
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from io import BytesIO
    from datetime import datetime, date
    from django.db.models import Q

    # Obtener parámetros de filtro
    fecha_str = request.GET.get('fecha')
    municipio_filtro = request.GET.get('municipio')

    if fecha_str:
        try:
            fecha_reporte = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except ValueError:
            fecha_reporte = None
    else:
        fecha_reporte = None

    # Obtener planillas según el rol del usuario
    if user.tipo_usuario == 'administrador':
        # Administradores ven todos los municipios
        planillas = Planilla.objects.all()
    elif user.tipo_usuario == 'tecnico':
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).distinct()
    elif user.tipo_usuario == 'vacunador':
        planillas = Planilla.objects.filter(
            Q(assigned_to=user) |
            Q(vacunadores_adicionales=user)
        ).distinct()
    else:
        planillas = Planilla.objects.none()

    # Filtrar por municipio si se proporciona
    if municipio_filtro:
        planillas = planillas.filter(municipio__icontains=municipio_filtro)

    # Obtener mascotas del municipio
    if user.tipo_usuario in ['administrador', 'tecnico']:
        # Administradores y técnicos ven todas las mascotas de sus municipios
        mascotas = Mascota.objects.filter(
            responsable__planilla__in=planillas
        ).select_related('responsable', 'responsable__planilla')
    else:
        # Vacunadores solo ven las mascotas que crearon
        mascotas = Mascota.objects.filter(
            created_by=user,
            responsable__planilla__in=planillas
        ).select_related('responsable', 'responsable__planilla')

    # Filtrar por fecha si se proporciona
    if fecha_reporte:
        mascotas = mascotas.filter(creado__date=fecha_reporte)

    # Agrupar mascotas por municipio, fecha y vacunador
    from collections import defaultdict
    mascotas_por_municipio_fecha_vacunador = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for mascota in mascotas.order_by('responsable__planilla__municipio', 'creado__date', 'created_by__username', 'responsable__nombre'):
        municipio = mascota.responsable.planilla.municipio
        fecha = mascota.creado.date()
        vacunador = mascota.created_by  # Usuario que creó la mascota
        mascotas_por_municipio_fecha_vacunador[municipio][fecha][vacunador].append(mascota)

    # Crear el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.5*inch, rightMargin=0.5*inch)

    # Contenido del PDF
    elements = []
    styles = getSampleStyleSheet()

    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=14,
        textColor=colors.darkblue,
        spaceAfter=12,
        alignment=1  # Centrado
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=10,
        alignment=1
    )

    # Si no hay mascotas
    if not mascotas_por_municipio_fecha_vacunador:
        title = Paragraph(
            f"Planilla de Vacunación - {user.get_tipo_usuario_display()}: {user.username}",
            title_style
        )
        elements.append(title)
        elements.append(Spacer(1, 20))

        no_data = Paragraph(
            "No hay mascotas vacunadas para los criterios seleccionados.",
            styles['Normal']
        )
        elements.append(no_data)
    else:
        # Procesar cada municipio
        first_page = True
        for municipio, fechas_dict in mascotas_por_municipio_fecha_vacunador.items():
            # Procesar cada fecha
            for fecha, vacunadores_dict in fechas_dict.items():
                # Procesar cada vacunador
                for vacunador, mascotas_list in vacunadores_dict.items():
                    if not first_page:
                        elements.append(PageBreak())
                    first_page = False

                    # Obtener información del vacunador
                    vacunador_nombre = vacunador.get_full_name() if vacunador and vacunador.get_full_name() else (vacunador.username if vacunador else "Sin asignar")
                    vacunador_username = vacunador.username if vacunador else "Sin asignar"

                    # Título de la página
                    title = Paragraph(
                        f"Planilla de Vacunación - {municipio}",
                        title_style
                    )
                    elements.append(title)

                    subtitle = Paragraph(
                        f"Fecha: {fecha.strftime('%d/%m/%Y')} | Vacunador: {vacunador_username} | Total: {len(mascotas_list)} mascotas",
                        subtitle_style
                    )
                    elements.append(subtitle)
                    elements.append(Spacer(1, 15))

                    # Dividir en páginas de 20 mascotas
                    MASCOTAS_POR_PAGINA = 20
                    total_paginas = (len(mascotas_list) + MASCOTAS_POR_PAGINA - 1) // MASCOTAS_POR_PAGINA

                    for pagina in range(total_paginas):
                        if pagina > 0:
                            elements.append(PageBreak())
                            # Repetir título en cada página
                            elements.append(Paragraph(
                                f"Planilla de Vacunación - {municipio} (Continuación)",
                                title_style
                            ))
                            elements.append(Paragraph(
                                f"Fecha: {fecha.strftime('%d/%m/%Y')} | Vacunador: {vacunador_username} | Página {pagina + 1} de {total_paginas}",
                                subtitle_style
                            ))
                            elements.append(Spacer(1, 15))

                        inicio = pagina * MASCOTAS_POR_PAGINA
                        fin = min(inicio + MASCOTAS_POR_PAGINA, len(mascotas_list))
                        mascotas_pagina = mascotas_list[inicio:fin]

                        # Crear tabla de mascotas
                        data = [[
                            '#',
                            'Responsable',
                            'Teléfono',
                            'Mascota',
                            'Tipo',
                            'Raza',
                            'Color',
                            'Antec.\nVacunal',
                            'Finca/Predio',
                            'Ubicación',
                            'Lote\nVacuna'
                        ]]

                        for i, mascota in enumerate(mascotas_pagina, inicio + 1):
                            responsable = mascota.responsable

                            # Determinar ubicación (vereda, centro poblado o barrio)
                            if responsable.zona == 'vereda':
                                ubicacion = f"V: {responsable.nombre_zona}"
                            elif responsable.zona == 'centro poblado':
                                ubicacion = f"CP: {responsable.nombre_zona}"
                            elif responsable.zona == 'barrio':
                                ubicacion = f"B: {responsable.nombre_zona}"
                            else:
                                ubicacion = responsable.nombre_zona

                            data.append([
                                str(i),
                                responsable.nombre[:20],  # Limitar caracteres
                                responsable.telefono,
                                mascota.nombre[:15],
                                'P' if mascota.tipo == 'perro' else 'G',
                                mascota.raza,
                                mascota.color[:12],
                                'Sí' if mascota.antecedente_vacunal else 'No',
                                responsable.finca[:15],
                                ubicacion[:20],
                                responsable.lote_vacuna[:10]
                            ])

                        # Configurar tabla
                        table = Table(data, colWidths=[
                            0.3*inch,  # #
                            1.2*inch,  # Responsable
                            0.8*inch,  # Teléfono
                            0.9*inch,  # Mascota
                            0.4*inch,  # Tipo
                            0.5*inch,  # Raza
                            0.8*inch,  # Color
                            0.5*inch,  # Antec. Vacunal
                            1.0*inch,  # Finca
                            1.2*inch,  # Ubicación
                            0.6*inch   # Lote Vacuna
                        ])

                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 7),
                            ('FONTSIZE', (0, 1), (-1, -1), 6),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                            ('TOPPADDING', (0, 0), (-1, 0), 8),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7f7f7')])
                        ]))

                        elements.append(table)
                        elements.append(Spacer(1, 10))

                        # Agregar información del vacunador que diligencio esta página
                        info_vacunador_style = ParagraphStyle(
                            'InfoVacunador',
                            parent=styles['Normal'],
                            fontSize=9,
                            textColor=colors.black,
                            alignment=0  # Alineación izquierda
                        )

                        diligenciado_text = f"<b>Diligenciado por:</b> {vacunador_nombre}"
                        diligenciado = Paragraph(diligenciado_text, info_vacunador_style)
                        elements.append(diligenciado)

                        digitador_text = f"<b>Digitador por:</b> {vacunador_username}"
                        digitador = Paragraph(digitador_text, info_vacunador_style)
                        elements.append(digitador)

                        # Información de página
                        if pagina < total_paginas - 1:
                            elements.append(Spacer(1, 10))
                            info_pagina = Paragraph(
                                f"<i>Mostrando mascotas {inicio + 1} a {fin} de {len(mascotas_list)}</i>",
                                subtitle_style
                            )
                            elements.append(info_pagina)

    # Pie de página general
    elements.append(Spacer(1, 20))
    footer_text = f"Reporte generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} - Sistema VetControl"
    footer = Paragraph(footer_text, subtitle_style)
    elements.append(footer)

    # Leyenda
    leyenda_text = "<i>Leyenda: P=Perro, G=Gato, V=Vereda, CP=Centro Poblado, B=Barrio</i>"
    leyenda = Paragraph(leyenda_text, subtitle_style)
    elements.append(leyenda)

    # Construir el PDF
    doc.build(elements)

    # Obtener el contenido del buffer
    pdf = buffer.getvalue()
    buffer.close()

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'planilla_{municipio_filtro or "todos"}_{fecha_reporte.strftime("%Y-%m-%d") if fecha_reporte else "todas_fechas"}_{user.username}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)

    return response

