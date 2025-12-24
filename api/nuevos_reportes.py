# Nuevos reportes PDF para VetControl
# IMPORTANTE: Importar pdf_utils ANTES que ReportLab para compatibilidad Python 3.8
from . import pdf_utils

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from django.db.models import Q, Count
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from io import BytesIO
from datetime import datetime, date
from collections import defaultdict
from .models import Responsable, Mascota, Planilla


@login_required
def reporte_municipio_por_dia_pdf(request):
    """
    Reporte 1: Por Municipio organizado por Día
    Muestra todos los registros agrupados por municipio y luego por día
    """
    user = request.user
    if user.tipo_usuario not in ['administrador', 'tecnico']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener el parámetro de municipio (si existe)
    municipio_filtro = request.GET.get('municipio', '').strip()

    # Obtener todas las mascotas con sus relaciones
    mascotas_query = Mascota.objects.select_related(
        'responsable__planilla', 'created_by'
    )

    # Filtrar según el tipo de usuario
    if user.tipo_usuario == 'tecnico':
        # Técnico: solo ve mascotas de sus planillas asignadas
        mascotas_query = mascotas_query.filter(
            Q(responsable__planilla__tecnico_asignado=user) |
            Q(responsable__planilla__tecnicos_adicionales=user)
        ).distinct()

    # Filtrar por municipio si se especifica
    if municipio_filtro:
        mascotas_query = mascotas_query.filter(
            responsable__planilla__municipio=municipio_filtro
        )

    mascotas = mascotas_query.order_by('responsable__planilla__municipio', 'creado')

    # Agrupar por municipio y luego por día
    data_por_municipio = defaultdict(lambda: defaultdict(list))

    for mascota in mascotas:
        municipio = mascota.responsable.planilla.municipio
        fecha = mascota.creado.date() if hasattr(mascota.creado, 'date') else mascota.creado
        data_por_municipio[municipio][fecha].append(mascota)

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12
    )

    # Título
    title = Paragraph("Reporte de Vacunación por Municipio (Organizado por Día)", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Diccionario para totales estadísticos por municipio
    estadisticas_por_municipio = {}

    # Procesar cada municipio
    for municipio in sorted(data_por_municipio.keys()):
        # Subtítulo del municipio
        municipio_title = Paragraph(f"<b>Municipio: {municipio}</b>", subtitle_style)
        elements.append(municipio_title)
        elements.append(Spacer(1, 0.1*inch))

        dias = data_por_municipio[municipio]

        # Tabla para este municipio - organizada por días
        table_data_mun = [[
            'Fecha',
            'Perros\nUrbano',
            'Perros\nRural',
            'Total\nPerros',
            'Gatos\nUrbano',
            'Gatos\nRural',
            'Total\nGatos',
            'Total\nUrbano',
            'Total\nRural',
            'TOTAL'
        ]]

        # Inicializar contadores para este municipio
        stats = {
            'perros_urbano': 0, 'perros_rural': 0,
            'gatos_urbano': 0, 'gatos_rural': 0
        }

        for fecha in sorted(dias.keys()):
            mascotas_dia = dias[fecha]

            # Contadores para este día
            stats_dia = {
                'perros_urbano': 0, 'perros_rural': 0,
                'gatos_urbano': 0, 'gatos_rural': 0
            }

            for mascota in mascotas_dia:
                resp = mascota.responsable

                # Contabilizar para estadísticas
                zona = (resp.zona or '').lower().strip()
                es_urbano = zona == 'barrio'
                es_rural = zona in ['vereda', 'centro poblado']

                if mascota.tipo == 'perro':
                    if es_urbano:
                        stats_dia['perros_urbano'] += 1
                        stats['perros_urbano'] += 1
                    elif es_rural:
                        stats_dia['perros_rural'] += 1
                        stats['perros_rural'] += 1
                elif mascota.tipo == 'gato':
                    if es_urbano:
                        stats_dia['gatos_urbano'] += 1
                        stats['gatos_urbano'] += 1
                    elif es_rural:
                        stats_dia['gatos_rural'] += 1
                        stats['gatos_rural'] += 1

            # Agregar fila para este día
            total_perros_dia = stats_dia['perros_urbano'] + stats_dia['perros_rural']
            total_gatos_dia = stats_dia['gatos_urbano'] + stats_dia['gatos_rural']
            total_urbano_dia = stats_dia['perros_urbano'] + stats_dia['gatos_urbano']
            total_rural_dia = stats_dia['perros_rural'] + stats_dia['gatos_rural']
            total_dia = total_perros_dia + total_gatos_dia

            table_data_mun.append([
                fecha.strftime('%d/%m/%Y'),
                str(stats_dia['perros_urbano']),
                str(stats_dia['perros_rural']),
                str(total_perros_dia),
                str(stats_dia['gatos_urbano']),
                str(stats_dia['gatos_rural']),
                str(total_gatos_dia),
                str(total_urbano_dia),
                str(total_rural_dia),
                str(total_dia)
            ])

        # Agregar subtotal del municipio
        total_perros = stats['perros_urbano'] + stats['perros_rural']
        total_gatos = stats['gatos_urbano'] + stats['gatos_rural']
        total_urbano = stats['perros_urbano'] + stats['gatos_urbano']
        total_rural = stats['perros_rural'] + stats['gatos_rural']
        total_general = total_perros + total_gatos

        table_data_mun.append([
            f'SUBTOTAL {municipio}',
            str(stats['perros_urbano']),
            str(stats['perros_rural']),
            str(total_perros),
            str(stats['gatos_urbano']),
            str(stats['gatos_rural']),
            str(total_gatos),
            str(total_urbano),
            str(total_rural),
            str(total_general)
        ])

        # Crear tabla para este municipio
        table = Table(table_data_mun, colWidths=[1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#90EE90')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Guardar estadísticas del municipio
        estadisticas_por_municipio[municipio] = stats

    # TOTALES GENERALES
    elements.append(PageBreak())
    elements.append(Paragraph("<b>TOTALES GENERALES</b>", title_style))
    elements.append(Spacer(1, 0.2*inch))

    totales = {
        'perros_urbano': 0, 'perros_rural': 0,
        'gatos_urbano': 0, 'gatos_rural': 0
    }

    for stats in estadisticas_por_municipio.values():
        totales['perros_urbano'] += stats['perros_urbano']
        totales['perros_rural'] += stats['perros_rural']
        totales['gatos_urbano'] += stats['gatos_urbano']
        totales['gatos_rural'] += stats['gatos_rural']

    total_perros_final = totales['perros_urbano'] + totales['perros_rural']
    total_gatos_final = totales['gatos_urbano'] + totales['gatos_rural']
    total_urbano_final = totales['perros_urbano'] + totales['gatos_urbano']
    total_rural_final = totales['perros_rural'] + totales['gatos_rural']
    total_final = total_perros_final + total_gatos_final

    totales_table_data = [
        ['Perros Urbano', 'Perros Rural', 'Total Perros', 'Gatos Urbano', 'Gatos Rural', 'Total Gatos', 'Total Urbano', 'Total Rural', 'TOTAL'],
        [str(totales['perros_urbano']), str(totales['perros_rural']), str(total_perros_final),
         str(totales['gatos_urbano']), str(totales['gatos_rural']), str(total_gatos_final),
         str(total_urbano_final), str(total_rural_final), str(total_final)]
    ]

    totales_table = Table(totales_table_data, colWidths=[0.9*inch] * 9)
    totales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffd700')),
        ('GRID', (0, 0), (-1, -1), 2, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
    ]))

    elements.append(totales_table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_municipio_por_dia_{datetime.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    return response


@login_required
def reporte_dia_por_municipio_pdf(request):
    """
    Reporte 2: Por Día organizado por Municipio
    Muestra todos los registros agrupados por día y luego por municipio
    """
    user = request.user
    if user.tipo_usuario not in ['administrador', 'tecnico']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener el parámetro de municipio (si existe)
    municipio_filtro = request.GET.get('municipio', '').strip()

    # Obtener todas las mascotas con sus relaciones
    mascotas_query = Mascota.objects.select_related(
        'responsable__planilla', 'created_by'
    )

    # Filtrar según el tipo de usuario
    if user.tipo_usuario == 'tecnico':
        # Técnico: solo ve mascotas de sus planillas asignadas
        mascotas_query = mascotas_query.filter(
            Q(responsable__planilla__tecnico_asignado=user) |
            Q(responsable__planilla__tecnicos_adicionales=user)
        ).distinct()

    # Filtrar por municipio si se especifica
    if municipio_filtro:
        mascotas_query = mascotas_query.filter(
            responsable__planilla__municipio=municipio_filtro
        )

    mascotas = mascotas_query.order_by('creado', 'responsable__planilla__municipio')

    # Agrupar por día y luego por municipio
    data_por_dia = defaultdict(lambda: defaultdict(list))

    for mascota in mascotas:
        fecha = mascota.creado.date() if hasattr(mascota.creado, 'date') else mascota.creado
        municipio = mascota.responsable.planilla.municipio
        data_por_dia[fecha][municipio].append(mascota)

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12
    )

    # Título
    title = Paragraph("Reporte de Vacunación por Día (Organizado por Municipio)", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Diccionario para totales estadísticos por día
    estadisticas_por_dia = {}

    # Procesar cada día
    for fecha in sorted(data_por_dia.keys()):
        # Subtítulo del día
        dia_title = Paragraph(f"<b>Fecha: {fecha.strftime('%d de %B de %Y')}</b>", subtitle_style)
        elements.append(dia_title)
        elements.append(Spacer(1, 0.1*inch))

        municipios = data_por_dia[fecha]

        # Tabla para este día - organizada por municipios
        table_data_dia = [[
            'Municipio',
            'Perros\nUrbano',
            'Perros\nRural',
            'Total\nPerros',
            'Gatos\nUrbano',
            'Gatos\nRural',
            'Total\nGatos',
            'Total\nUrbano',
            'Total\nRural',
            'TOTAL'
        ]]

        # Inicializar contadores para este día
        stats = {
            'perros_urbano': 0, 'perros_rural': 0,
            'gatos_urbano': 0, 'gatos_rural': 0
        }

        for municipio in sorted(municipios.keys()):
            mascotas_municipio = municipios[municipio]

            # Contadores para este municipio en este día
            stats_mun = {
                'perros_urbano': 0, 'perros_rural': 0,
                'gatos_urbano': 0, 'gatos_rural': 0
            }

            for mascota in mascotas_municipio:
                resp = mascota.responsable

                # Contabilizar para estadísticas
                zona = (resp.zona or '').lower().strip()
                es_urbano = zona == 'barrio'
                es_rural = zona in ['vereda', 'centro poblado']

                if mascota.tipo == 'perro':
                    if es_urbano:
                        stats_mun['perros_urbano'] += 1
                        stats['perros_urbano'] += 1
                    elif es_rural:
                        stats_mun['perros_rural'] += 1
                        stats['perros_rural'] += 1
                elif mascota.tipo == 'gato':
                    if es_urbano:
                        stats_mun['gatos_urbano'] += 1
                        stats['gatos_urbano'] += 1
                    elif es_rural:
                        stats_mun['gatos_rural'] += 1
                        stats['gatos_rural'] += 1

            # Agregar fila para este municipio
            total_perros_mun = stats_mun['perros_urbano'] + stats_mun['perros_rural']
            total_gatos_mun = stats_mun['gatos_urbano'] + stats_mun['gatos_rural']
            total_urbano_mun = stats_mun['perros_urbano'] + stats_mun['gatos_urbano']
            total_rural_mun = stats_mun['perros_rural'] + stats_mun['gatos_rural']
            total_mun = total_perros_mun + total_gatos_mun

            table_data_dia.append([
                municipio,
                str(stats_mun['perros_urbano']),
                str(stats_mun['perros_rural']),
                str(total_perros_mun),
                str(stats_mun['gatos_urbano']),
                str(stats_mun['gatos_rural']),
                str(total_gatos_mun),
                str(total_urbano_mun),
                str(total_rural_mun),
                str(total_mun)
            ])

        # Agregar subtotal del día
        total_perros = stats['perros_urbano'] + stats['perros_rural']
        total_gatos = stats['gatos_urbano'] + stats['gatos_rural']
        total_urbano = stats['perros_urbano'] + stats['gatos_urbano']
        total_rural = stats['perros_rural'] + stats['gatos_rural']
        total_general = total_perros + total_gatos

        table_data_dia.append([
            f'SUBTOTAL {fecha.strftime("%d/%m/%Y")}',
            str(stats['perros_urbano']),
            str(stats['perros_rural']),
            str(total_perros),
            str(stats['gatos_urbano']),
            str(stats['gatos_rural']),
            str(total_gatos),
            str(total_urbano),
            str(total_rural),
            str(total_general)
        ])

        # Crear tabla para este día
        table = Table(table_data_dia, colWidths=[1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#90EE90')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Guardar estadísticas del día
        estadisticas_por_dia[fecha] = stats

    # TOTALES GENERALES
    elements.append(PageBreak())
    elements.append(Paragraph("<b>TOTALES GENERALES</b>", title_style))
    elements.append(Spacer(1, 0.2*inch))

    totales = {
        'perros_urbano': 0, 'perros_rural': 0,
        'gatos_urbano': 0, 'gatos_rural': 0
    }

    for stats in estadisticas_por_dia.values():
        totales['perros_urbano'] += stats['perros_urbano']
        totales['perros_rural'] += stats['perros_rural']
        totales['gatos_urbano'] += stats['gatos_urbano']
        totales['gatos_rural'] += stats['gatos_rural']

    total_perros_final = totales['perros_urbano'] + totales['perros_rural']
    total_gatos_final = totales['gatos_urbano'] + totales['gatos_rural']
    total_urbano_final = totales['perros_urbano'] + totales['gatos_urbano']
    total_rural_final = totales['perros_rural'] + totales['gatos_rural']
    total_final = total_perros_final + total_gatos_final

    totales_table_data = [
        ['Perros Urbano', 'Perros Rural', 'Total Perros', 'Gatos Urbano', 'Gatos Rural', 'Total Gatos', 'Total Urbano', 'Total Rural', 'TOTAL'],
        [str(totales['perros_urbano']), str(totales['perros_rural']), str(total_perros_final),
         str(totales['gatos_urbano']), str(totales['gatos_rural']), str(total_gatos_final),
         str(total_urbano_final), str(total_rural_final), str(total_final)]
    ]

    totales_table = Table(totales_table_data, colWidths=[0.9*inch] * 9)
    totales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffd700')),
        ('GRID', (0, 0), (-1, -1), 2, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
    ]))

    elements.append(totales_table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_dia_por_municipio_{datetime.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    return response


@login_required
def reporte_estadistico_rango_fechas_pdf(request):
    """
    Reporte 3: Estadístico por Rango de Fechas
    Muestra totales de perros/gatos urbano/rural por municipio en un rango de fechas
    """
    user = request.user
    if user.tipo_usuario not in ['administrador', 'tecnico']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener parámetros de fecha
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Parsear fechas
    try:
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        else:
            fecha_inicio = None

        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        else:
            fecha_fin = None
    except ValueError:
        messages.error(request, 'Formato de fecha inválido. Use AAAA-MM-DD')
        return redirect('dashboard_administrador')

    # Filtrar mascotas por rango de fechas (si se especifican)
    mascotas_query = Mascota.objects.select_related('responsable__planilla')

    # Filtrar según el tipo de usuario
    if user.tipo_usuario == 'tecnico':
        # Técnico: solo ve mascotas de sus planillas asignadas
        mascotas_query = mascotas_query.filter(
            Q(responsable__planilla__tecnico_asignado=user) |
            Q(responsable__planilla__tecnicos_adicionales=user)
        ).distinct()

    if fecha_inicio and fecha_fin:
        # Si se especifican ambas fechas
        mascotas = mascotas_query.filter(
            creado__date__gte=fecha_inicio,
            creado__date__lte=fecha_fin
        )
    elif fecha_inicio:
        # Solo fecha de inicio
        mascotas = mascotas_query.filter(creado__date__gte=fecha_inicio)
        fecha_fin = date.today()
    elif fecha_fin:
        # Solo fecha fin
        mascotas = mascotas_query.filter(creado__date__lte=fecha_fin)
        # Obtener la fecha más antigua de los registros
        primera_mascota = mascotas_query.order_by('creado').first()
        fecha_inicio = primera_mascota.creado.date() if primera_mascota else date.today()
    else:
        # Sin filtro de fechas - TODOS los registros
        mascotas = mascotas_query.all()
        # Obtener rango completo
        primera_mascota = mascotas_query.order_by('creado').first()
        ultima_mascota = mascotas_query.order_by('-creado').first()
        fecha_inicio = primera_mascota.creado.date() if primera_mascota else date.today()
        fecha_fin = ultima_mascota.creado.date() if ultima_mascota else date.today()

    # Agrupar por municipio y calcular estadísticas
    stats_por_municipio = defaultdict(lambda: {
        'perros_urbano': 0, 'perros_rural': 0, 'perros_total': 0,
        'gatos_urbano': 0, 'gatos_rural': 0, 'gatos_total': 0,
        'total_urbano': 0, 'total_rural': 0, 'total': 0
    })

    for mascota in mascotas:
        municipio = mascota.responsable.planilla.municipio
        zona = (mascota.responsable.zona or '').lower().strip()
        tipo = mascota.tipo

        # Determinar si es urbano o rural (case-insensitive)
        # URBANO: solo barrio
        # RURAL: vereda y centro poblado
        es_urbano = zona == 'barrio'
        es_rural = zona in ['vereda', 'centro poblado']

        # Contadores
        if tipo == 'perro':
            stats_por_municipio[municipio]['perros_total'] += 1
            if es_urbano:
                stats_por_municipio[municipio]['perros_urbano'] += 1
            elif es_rural:
                stats_por_municipio[municipio]['perros_rural'] += 1
        else:  # gato
            stats_por_municipio[municipio]['gatos_total'] += 1
            if es_urbano:
                stats_por_municipio[municipio]['gatos_urbano'] += 1
            elif es_rural:
                stats_por_municipio[municipio]['gatos_rural'] += 1

        if es_urbano:
            stats_por_municipio[municipio]['total_urbano'] += 1
        elif es_rural:
            stats_por_municipio[municipio]['total_rural'] += 1

        stats_por_municipio[municipio]['total'] += 1

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30,
                          topMargin=30, bottomMargin=18)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=1
    )

    # Título
    title = Paragraph(
        f"Reporte Estadístico de Vacunación<br/>"
        f"<font size=12>Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}</font>",
        title_style
    )
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Tabla de estadísticas
    table_data = [[
        'Municipio',
        'Perros\nUrbano',
        'Perros\nRural',
        'Total\nPerros',
        'Gatos\nUrbano',
        'Gatos\nRural',
        'Total\nGatos',
        'Total\nUrbano',
        'Total\nRural',
        'TOTAL'
    ]]

    # Totales generales
    totales = {
        'perros_urbano': 0, 'perros_rural': 0, 'perros_total': 0,
        'gatos_urbano': 0, 'gatos_rural': 0, 'gatos_total': 0,
        'total_urbano': 0, 'total_rural': 0, 'total': 0
    }

    # Agregar datos por municipio
    for municipio in sorted(stats_por_municipio.keys()):
        stats = stats_por_municipio[municipio]
        table_data.append([
            municipio,
            str(stats['perros_urbano']),
            str(stats['perros_rural']),
            str(stats['perros_total']),
            str(stats['gatos_urbano']),
            str(stats['gatos_rural']),
            str(stats['gatos_total']),
            str(stats['total_urbano']),
            str(stats['total_rural']),
            str(stats['total'])
        ])

        # Sumar a totales
        for key in totales:
            totales[key] += stats[key]

    # Agregar fila de totales
    table_data.append([
        'TOTALES',
        str(totales['perros_urbano']),
        str(totales['perros_rural']),
        str(totales['perros_total']),
        str(totales['gatos_urbano']),
        str(totales['gatos_rural']),
        str(totales['gatos_total']),
        str(totales['total_urbano']),
        str(totales['total_rural']),
        str(totales['total'])
    ])

    table = Table(table_data, colWidths=[1.5*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a90e2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ffd700')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey])
    ]))

    elements.append(table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_estadistico_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    return response


@login_required
def reporte_listado_lugares_pdf(request):
    """
    Reporte: Listado de Lugares (Veredas, Barrios, Centros Poblados)
    Muestra todos los lugares con la cantidad de mascotas vacunadas en cada uno
    Disponible para técnicos y administradores
    """
    user = request.user
    if user.tipo_usuario not in ['administrador', 'tecnico']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener parámetro de municipio
    municipio_filtro = request.GET.get('municipio', '').strip()

    # Obtener planillas según el rol del usuario
    if user.tipo_usuario == 'administrador':
        planillas = Planilla.objects.all()
    else:  # tecnico
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).distinct()

    # Filtrar por municipio si se proporciona
    if municipio_filtro:
        planillas = planillas.filter(municipio__icontains=municipio_filtro)

    # Obtener responsables de las planillas
    responsables = Responsable.objects.filter(
        planilla__in=planillas
    ).select_related('planilla').prefetch_related('mascotas')

    # Estructura: {municipio: {nombre_zona: cantidad}}
    datos_por_municipio = defaultdict(lambda: defaultdict(int))

    # Procesar cada responsable y contar mascotas por lugar
    for responsable in responsables:
        municipio = responsable.planilla.municipio

        # Manejar zonas vacías - priorizar nombre_zona sobre zona
        nombre_zona = None
        if responsable.nombre_zona and responsable.nombre_zona.strip():
            nombre_zona = responsable.nombre_zona.strip()
        elif responsable.zona and responsable.zona.strip():
            nombre_zona = responsable.zona.strip()
        else:
            nombre_zona = 'Sin especificar'

        # Contar todas las mascotas del responsable
        cantidad_mascotas = responsable.mascotas.count()

        # Solo agregar si hay mascotas
        if cantidad_mascotas > 0:
            datos_por_municipio[municipio][nombre_zona] += cantidad_mascotas

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo_texto = f'Listado de Lugares - {municipio_filtro if municipio_filtro else "Todos los Municipios"}'
    titulo = Paragraph(titulo_texto, ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=1
    ))
    elements.append(titulo)

    # Información del reporte
    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    info = Paragraph(
        f'Generado: {fecha_actual} | Usuario: {user.username}',
        ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=1
        )
    )
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))

    # Si no hay datos
    if not datos_por_municipio:
        mensaje = Paragraph(
            'No se encontraron datos para los criterios seleccionados.',
            ParagraphStyle(
                'NoData',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.red,
                alignment=1
            )
        )
        elements.append(mensaje)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="listado_lugares_{municipio_filtro or "todos"}.pdf"'
        response.write(pdf)
        return response

    # Procesar cada municipio
    for municipio in sorted(datos_por_municipio.keys()):
        # Título del municipio
        titulo_municipio = Paragraph(
            f'<b>Municipio: {municipio}</b>',
            ParagraphStyle(
                'MunicipioTitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#34495e'),
                spaceAfter=10
            )
        )
        elements.append(titulo_municipio)

        # Preparar datos de la tabla
        lugares_data = datos_por_municipio[municipio]

        # Crear lista de datos ordenada
        lugares_ordenados = sorted(lugares_data.items(), key=lambda x: x[1], reverse=True)

        # Calcular totales
        total_mascotas = sum(lugares_data.values())
        total_lugares = len(lugares_data)

        # Construir tabla
        table_data = [
            ['Lugar', 'Tipo', 'Mascotas Vacunadas']
        ]

        for nombre_zona, cantidad in lugares_ordenados:
            # Determinar tipo de lugar
            nombre_lower = nombre_zona.lower()
            if any(x in nombre_lower for x in ['vereda', 'vda', 'v.']):
                tipo = 'Vereda'
            elif any(x in nombre_lower for x in ['barrio', 'br', 'b.']):
                tipo = 'Barrio'
            elif any(x in nombre_lower for x in ['centro poblado', 'cp', 'c.p']):
                tipo = 'Centro Poblado'
            else:
                tipo = 'Otro'

            table_data.append([
                nombre_zona,
                tipo,
                str(cantidad)
            ])

        # Agregar fila de totales
        table_data.append([
            f'TOTAL: {total_lugares} lugares',
            '',
            f'{total_mascotas}'
        ])

        # Crear tabla
        table = Table(table_data, colWidths=[3.5*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Contenido
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ALIGN', (0, 1), (0, -2), 'LEFT'),
            ('ALIGN', (1, 1), (1, -2), 'CENTER'),
            ('ALIGN', (2, 1), (-1, -2), 'CENTER'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),

            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            ('GRID', (0, -1), (-1, -1), 1, colors.HexColor('#3498db')),

            # Alternancia de colores
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')])
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.4*inch))

    # Si hay múltiples municipios, agregar resumen general
    if len(datos_por_municipio) > 1:
        elements.append(PageBreak())

        # Título resumen
        titulo_resumen = Paragraph(
            '<b>Resumen General</b>',
            ParagraphStyle(
                'ResumenTitle',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#34495e'),
                spaceAfter=10
            )
        )
        elements.append(titulo_resumen)

        # Datos del resumen
        resumen_data = [['Municipio', 'Total Lugares', 'Total Mascotas']]

        total_general_lugares = 0
        total_general_mascotas = 0

        for municipio in sorted(datos_por_municipio.keys()):
            lugares = len(datos_por_municipio[municipio])
            mascotas = sum(datos_por_municipio[municipio].values())
            total_general_lugares += lugares
            total_general_mascotas += mascotas

            resumen_data.append([
                municipio,
                str(lugares),
                str(mascotas)
            ])

        # Totales generales
        resumen_data.append([
            'TOTAL GENERAL',
            str(total_general_lugares),
            str(total_general_mascotas)
        ])

        # Crear tabla de resumen
        resumen_table = Table(resumen_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        resumen_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Contenido
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (0, 1), (0, -2), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -2), 'CENTER'),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),

            # Fila de totales
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            ('GRID', (0, -1), (-1, -1), 1, colors.HexColor('#27ae60')),

            # Alternancia de colores
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#e8f8f5')])
        ]))

        elements.append(resumen_table)

    # Construir PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="listado_lugares_{municipio_filtro or "todos"}_{user.username}.pdf"'
    response.write(pdf)
    return response


@login_required
def reporte_listado_lugares_excel(request):
    """
    Reporte: Listado de Lugares en Excel
    Exporta el listado de lugares con cantidad de mascotas vacunadas a Excel
    Disponible para técnicos y administradores
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    user = request.user
    if user.tipo_usuario not in ['administrador', 'tecnico']:
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener parámetro de municipio
    municipio_filtro = request.GET.get('municipio', '').strip()

    # Obtener planillas según el rol del usuario
    if user.tipo_usuario == 'administrador':
        planillas = Planilla.objects.all()
    else:  # tecnico
        planillas = Planilla.objects.filter(
            Q(tecnico_asignado=user) |
            Q(tecnicos_adicionales=user)
        ).distinct()

    # Filtrar por municipio si se proporciona
    if municipio_filtro:
        planillas = planillas.filter(municipio__icontains=municipio_filtro)

    # Obtener responsables de las planillas
    responsables = Responsable.objects.filter(
        planilla__in=planillas
    ).select_related('planilla').prefetch_related('mascotas')

    # Estructura: {municipio: {nombre_zona: cantidad}}
    datos_por_municipio = defaultdict(lambda: defaultdict(int))

    # Procesar cada responsable y contar mascotas por lugar
    for responsable in responsables:
        municipio = responsable.planilla.municipio

        # Manejar zonas vacías - priorizar nombre_zona sobre zona
        nombre_zona = None
        if responsable.nombre_zona and responsable.nombre_zona.strip():
            nombre_zona = responsable.nombre_zona.strip()
        elif responsable.zona and responsable.zona.strip():
            nombre_zona = responsable.zona.strip()
        else:
            nombre_zona = 'Sin especificar'

        # Contar todas las mascotas del responsable
        cantidad_mascotas = responsable.mascotas.count()

        # Solo agregar si hay mascotas
        if cantidad_mascotas > 0:
            datos_por_municipio[municipio][nombre_zona] += cantidad_mascotas

    # Crear workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado de Lugares"

    # Estilos
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    municipio_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    municipio_font = Font(bold=True, color="FFFFFF", size=11)

    total_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    total_font = Font(bold=True, size=10)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:C1')
    titulo_cell = ws['A1']
    titulo_cell.value = f'Listado de Lugares - {municipio_filtro if municipio_filtro else "Todos los Municipios"}'
    titulo_cell.font = Font(bold=True, size=14)
    titulo_cell.alignment = Alignment(horizontal="center")

    # Información
    ws.merge_cells('A2:C2')
    info_cell = ws['A2']
    info_cell.value = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Usuario: {user.username}'
    info_cell.alignment = Alignment(horizontal="center")

    current_row = 4

    # Si no hay datos
    if not datos_por_municipio:
        ws['A4'] = 'No se encontraron datos para los criterios seleccionados.'
        ws['A4'].font = Font(color="FF0000", size=12)
        ws['A4'].alignment = Alignment(horizontal="center")
    else:
        # Procesar cada municipio
        for municipio in sorted(datos_por_municipio.keys()):
            # Título del municipio
            ws.merge_cells(f'A{current_row}:C{current_row}')
            municipio_cell = ws[f'A{current_row}']
            municipio_cell.value = f'Municipio: {municipio}'
            municipio_cell.fill = municipio_fill
            municipio_cell.font = municipio_font
            municipio_cell.alignment = Alignment(horizontal="center")
            current_row += 1

            # Encabezados de columnas
            headers = ['Lugar', 'Tipo', 'Mascotas Vacunadas']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            current_row += 1

            # Datos de lugares
            lugares_data = datos_por_municipio[municipio]
            lugares_ordenados = sorted(lugares_data.items(), key=lambda x: x[1], reverse=True)

            total_mascotas = 0
            for nombre_zona, cantidad in lugares_ordenados:
                # Determinar tipo de lugar
                nombre_lower = nombre_zona.lower()
                if any(x in nombre_lower for x in ['vereda', 'vda', 'v.']):
                    tipo = 'Vereda'
                elif any(x in nombre_lower for x in ['barrio', 'br', 'b.']):
                    tipo = 'Barrio'
                elif any(x in nombre_lower for x in ['centro poblado', 'cp', 'c.p']):
                    tipo = 'Centro Poblado'
                else:
                    tipo = 'Otro'

                # Escribir fila
                ws.cell(row=current_row, column=1, value=nombre_zona).border = border
                ws.cell(row=current_row, column=2, value=tipo).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=2).border = border
                ws.cell(row=current_row, column=3, value=cantidad).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=3).border = border

                total_mascotas += cantidad
                current_row += 1

            # Fila de totales
            ws.cell(row=current_row, column=1, value=f'TOTAL: {len(lugares_ordenados)} lugares').fill = total_fill
            ws.cell(row=current_row, column=1).font = total_font
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=2).fill = total_fill
            ws.cell(row=current_row, column=2).border = border
            ws.cell(row=current_row, column=3, value=total_mascotas).fill = total_fill
            ws.cell(row=current_row, column=3).font = total_font
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3).border = border

            current_row += 3  # Espacio entre municipios

        # Si hay múltiples municipios, agregar resumen
        if len(datos_por_municipio) > 1:
            current_row += 1
            ws.merge_cells(f'A{current_row}:C{current_row}')
            resumen_cell = ws[f'A{current_row}']
            resumen_cell.value = 'RESUMEN GENERAL'
            resumen_cell.font = Font(bold=True, size=13)
            resumen_cell.alignment = Alignment(horizontal="center")
            current_row += 1

            # Encabezados resumen
            headers_resumen = ['Municipio', 'Total Lugares', 'Total Mascotas']
            for col_num, header in enumerate(headers_resumen, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = header_alignment
                cell.border = border
            current_row += 1

            # Datos resumen
            total_general_lugares = 0
            total_general_mascotas = 0

            for municipio in sorted(datos_por_municipio.keys()):
                lugares = len(datos_por_municipio[municipio])
                mascotas = sum(datos_por_municipio[municipio].values())
                total_general_lugares += lugares
                total_general_mascotas += mascotas

                ws.cell(row=current_row, column=1, value=municipio).border = border
                ws.cell(row=current_row, column=2, value=lugares).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=2).border = border
                ws.cell(row=current_row, column=3, value=mascotas).alignment = Alignment(horizontal="center")
                ws.cell(row=current_row, column=3).border = border
                current_row += 1

            # Total general
            ws.cell(row=current_row, column=1, value='TOTAL GENERAL').fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=2, value=total_general_lugares).fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
            ws.cell(row=current_row, column=2).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=2).border = border
            ws.cell(row=current_row, column=3, value=total_general_mascotas).fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
            ws.cell(row=current_row, column=3).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3).border = border

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="listado_lugares_{municipio_filtro or "todos"}_{user.username}.xlsx"'
    return response


@login_required
def exportar_planillas_completas_excel(request):
    """
    Exporta toda la información de mascotas y responsables a Excel.
    OPTIMIZADO para 80,000+ registros usando write_only mode.
    """
    import openpyxl
    from openpyxl import Workbook

    user = request.user
    if user.tipo_usuario != 'administrador':
        messages.error(request, 'No tienes permisos para acceder a esta sección.')
        return redirect('login')

    # Obtener parámetro de municipio
    municipio_filtro = request.GET.get('municipio', '').strip()

    # Query base optimizada - solo campos necesarios
    mascotas_query = Mascota.objects.select_related(
        'responsable__planilla',
        'responsable__created_by',
        'created_by'
    ).defer('foto')  # Excluir campo de imagen

    # Filtrar por municipio si se especifica
    if municipio_filtro:
        mascotas_query = mascotas_query.filter(
            responsable__planilla__municipio=municipio_filtro
        )

    # Contar total
    total_registros = mascotas_query.count()

    # Crear workbook en modo write_only (mucho más rápido para grandes volúmenes)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Planillas Completas")

    # Encabezados
    headers = [
        'Municipio', 'Urbano/Rural', 'Tipo Zona', 'Nombre Lugar',
        'Nombre Responsable', 'Teléfono', 'Finca/Predio', 'Lote Vacuna',
        'Creado por (Resp)', 'Fecha Creación Resp',
        'Nombre Mascota', 'Tipo', 'Raza', 'Color',
        'Antecedente Vacunal', 'Esterilizado', 'Latitud', 'Longitud',
        'Creado por (Mascota)', 'Fecha Creación Mascota'
    ]

    # Escribir título
    ws.append([f'Planillas Completas - {municipio_filtro or "Todos los Municipios"}'])
    ws.append([f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")} | Usuario: {user.username} | Total: {total_registros}'])
    ws.append([])  # Fila vacía
    ws.append(headers)

    # Procesar en lotes de 1000 para eficiencia
    mascotas = mascotas_query.order_by(
        'responsable__planilla__municipio',
        'responsable__nombre',
        'nombre'
    ).iterator(chunk_size=1000)

    for mascota in mascotas:
        resp = mascota.responsable
        planilla = resp.planilla if resp else None

        row = [
            planilla.municipio if planilla else '',
            planilla.get_urbano_rural_display() if planilla else '',
            (resp.zona or '').title() if resp else '',
            resp.nombre_zona if resp else '',
            resp.nombre if resp else '',
            resp.telefono if resp else '',
            resp.finca if resp else '',
            resp.lote_vacuna if resp else '',
            resp.created_by.username if resp and resp.created_by else '',
            resp.creado.strftime('%d/%m/%Y %H:%M') if resp and resp.creado else '',
            mascota.nombre,
            mascota.tipo,
            mascota.raza,
            mascota.color,
            'Sí' if mascota.antecedente_vacunal else 'No',
            'Sí' if mascota.esterilizado else 'No',
            str(mascota.latitud) if mascota.latitud else '',
            str(mascota.longitud) if mascota.longitud else '',
            mascota.created_by.username if mascota.created_by else '',
            mascota.creado.strftime('%d/%m/%Y %H:%M') if mascota.creado else ''
        ]
        ws.append(row)

    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Nombre del archivo
    nombre_archivo = f'planillas_completas_{municipio_filtro or "todos"}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response
